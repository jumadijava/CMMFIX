"""
pages/report.py
───────────────
Report Hub — landing page dengan pilihan jenis laporan.
  • WSIRD Produksi  → QCL Inspection (format dari settings.py)
  • QIS Report      → coming soon
"""
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
from datetime import datetime
import io
import copy
import json

from local_db import save_report, get_reports, load_report_data, get_report_meta, send_ng_notif, update_report_status


def _send_ng_notifs_from_report(report: dict, username: str) -> int:
    """Kirim notifikasi untuk setiap item NG di laporan ke Produksi."""
    items        = report.get('items', [])
    measurements = report.get('measurements', {})
    samples      = report.get('sample_order', report.get('samples', []))
    header       = report.get('header', {})

    model_name = header.get('modelName', '')
    part_name  = header.get('partName', header.get('namaPart', ''))  # partName = PartName murni

    # Konversi tanggal ke format "%d %b %Y" agar match rc_key di diagnostic
    from datetime import datetime as _dt
    tgl_raw = header.get('tanggal', '')
    try:
        date_str = _dt.strptime(tgl_raw, '%d/%m/%Y').strftime('%d %b %Y')
    except Exception:
        date_str = tgl_raw
    shift_str = str(header.get('shift', ''))

    count = 0
    for item in items:
        iid  = item['id']
        meas = measurements.get(iid, {})
        for ni, sname in enumerate(samples):
            val_str = meas.get(f'n{ni+1}', '')
            if not val_str:
                continue
            if is_oot(val_str, item):
                send_ng_notif({
                    "from_user":   username,
                    "from_role":   "Measurement",
                    "to_role":     "Produksi",
                    "part":        part_name,
                    "model":       model_name,
                    "ref":         item.get('g', item.get('id','')),
                    "parameter":   item.get('nm', ''),
                    "sampleno":    sname,
                    "date":        date_str,
                    "shift":       shift_str,
                    "deviation":   val_str,
                    "kp":          "1" if item.get('kp') else "0",
                    "category":    "",
                    "description": "",
                    "pic":         "",
                    "status":      "Open",
                    "report_id":   report.get('id', ''),
                })
                count += 1
    return count

# PDF export via win32com (Windows + Excel required)
_WIN32_ERR = None
try:
    import win32com.client
    import pythoncom
    HAS_WIN32 = True
except Exception as _e:
    HAS_WIN32   = False
    _WIN32_ERR  = str(_e)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─────────────────────────────────────────────────────────────────────────────
# DATA — 110 TITIK INSPEKSI
# id, g=group, nm=nama, t=tools, nom=nominal, tp=tol+, tm=tol-, kp=kritikal, mv=max_val
# ─────────────────────────────────────────────────────────────────────────────
ITEMS = [
    # PAGE 1
    dict(id='B000',  g='B000',   nm='FLATNESS FACE CRANK CASE',       t='CMM',           nom='',       tp=None,  tm=None,  kp=False, mv=0.1 ),
    dict(id='A001a', g='A000-1', nm='KERATAAN SURFACE A',              t='CMM',           nom='',       tp=None,  tm=None,  kp=False, mv=0.05),
    dict(id='A001b', g='A000-1', nm='KERATAAN ADJACENT BOSSES',        t='CMM',           nom='',       tp=None,  tm=None,  kp=False, mv=0.03),
    dict(id='A002a', g='A000-2', nm='KERATAAN SURFACE B',              t='CMM',           nom='',       tp=None,  tm=None,  kp=False, mv=0.05),
    dict(id='A002b', g='A000-2', nm='KERATAAN ADJACENT BOSSES',        t='CMM',           nom='',       tp=None,  tm=None,  kp=False, mv=0.03),
    dict(id='A003a', g='A001',   nm='POSISI X',                        t='CMM',           nom='83',     tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A003b', g='A001',   nm='POSISI Y',                        t='CMM',           nom='110',    tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A003c', g='A001',   nm='POSISI X Tap',                    t='CMM',           nom='83',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A003d', g='A001',   nm='POSISI Y Tap',                    t='CMM',           nom='110',    tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A004a', g='A002',   nm='DISTANCE X FROM A001',            t='CMM',           nom='136',    tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A004b', g='A002',   nm='DISTANCE Y FROM A001',            t='CMM',           nom='163.5',  tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A004c', g='A002',   nm='DISTANCE X FROM A001 Tap',        t='CMM',           nom='136',    tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A004d', g='A002',   nm='DISTANCE Y FROM A001 Tap',        t='CMM',           nom='163.5',  tp=0.20,  tm=0.20,  kp=False, mv=None),
    # PAGE 2
    dict(id='A005a', g='A003',   nm='POSISI X DARI A006',              t='CMM',           nom='48.5',   tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A005b', g='A003',   nm='POSISI Y DARI A006',              t='CMM',           nom='20',     tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A006a', g='A004',   nm='POSISI X DARI A003',              t='CMM',           nom='112',    tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A006b', g='A004',   nm='POSISI Y DARI A003',              t='CMM',           nom='112.5',  tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A006c', g='A004',   nm='KETEGAKLURUSAN THD R',            t='CMM',           nom='Max 0.01/10', tp=None, tm=None, kp=False, mv=None),
    dict(id='A007a', g='A005',   nm='POSISI X',                        t='CMM',           nom='0',      tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A007b', g='A005',   nm='POSISI Y',                        t='CMM',           nom='0',      tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A007c', g='A005',   nm='KESUMBUAN DIA 45 THD S',          t='CMM',           nom='',       tp=None,  tm=None,  kp=True,  mv=0.03),
    dict(id='A008a', g='A006',   nm='POSISI X',                        t='CMM',           nom='258',    tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A008b', g='A006',   nm='POSISI Y',                        t='CMM',           nom='0',      tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A008c', g='A006',   nm='KESUMBUAN THD T',                 t='CMM',           nom='',       tp=None,  tm=None,  kp=False, mv=0.03),
    dict(id='A009a', g='A007',   nm='JARAK DARI A006',                 t='CMM',           nom='47',     tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A009b', g='A007',   nm='POSISI Y',                        t='CMM',           nom='0',      tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A010a', g='A008',   nm='POSISI X',                        t='CMM',           nom='62.5',   tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A010b', g='A008',   nm='POSISI Y',                        t='CMM',           nom='22.9',   tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='B011a', g='B011',   nm='POSISI X',                        t='CMM',           nom='62.5',   tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='B011b', g='B011',   nm='POSISI Y',                        t='CMM',           nom='43',     tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='B012a', g='B012',   nm='POSISI X DARI B011',              t='CMM',           nom='405.5',  tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='B012b', g='B012',   nm='POSISI Y DARI B011',              t='CMM',           nom='45.5',   tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='B012c', g='B012',   nm='POSISI X DARI B011 Tap',          t='CMM',           nom='405.5',  tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B012d', g='B012',   nm='POSISI Y DARI B011 Tap',          t='CMM',           nom='45.5',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A019a', g='A019',   nm='POSISI X',                        t='CMM',           nom='93',     tp=0.1,   tm=0.1,   kp=False, mv=None),
    dict(id='A019b', g='A019',   nm='POSISI Y',                        t='CMM',           nom='80',     tp=0.1,   tm=0.1,   kp=False, mv=None),
    dict(id='A020a', g='A020',   nm='POSISI X',                        t='CMM',           nom='243.5',  tp=0.05,  tm=0.05,  kp=False, mv=None),
    dict(id='A020b', g='A020',   nm='POSISI Y',                        t='CMM',           nom='51',     tp=0.05,  tm=0.05,  kp=False, mv=None),
    # PAGE 3
    dict(id='B021a', g='B021',   nm='POSISI X',                        t='CMM',           nom='126.5',  tp=0.20,  tm=0.20,  kp=True,  mv=None),
    dict(id='B021b', g='B021',   nm='POSISI Y',                        t='CMM',           nom='129',    tp=0.20,  tm=0.20,  kp=True,  mv=None),
    dict(id='B022a', g='B022',   nm='POSISI X',                        t='CMM',           nom='332.5',  tp=0.20,  tm=0.20,  kp=True,  mv=None),
    dict(id='B022b', g='B022',   nm='POSISI Y',                        t='CMM',           nom='119.8',  tp=0.20,  tm=0.20,  kp=True,  mv=None),
    dict(id='B023a', g='B023',   nm='POSISI X',                        t='CMM',           nom='68',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B023b', g='B023',   nm='POSISI Y',                        t='CMM',           nom='120.5',  tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B024a', g='B024',   nm='POSISI X',                        t='CMM',           nom='257',    tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B024b', g='B024',   nm='POSISI Y',                        t='CMM',           nom='95.5',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B025a', g='B025',   nm='POSISI X',                        t='CMM',           nom='13',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B025b', g='B025',   nm='POSISI Y',                        t='CMM',           nom='104',    tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B026a', g='B026',   nm='POSISI X',                        t='CMM',           nom='13',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B026b', g='B026',   nm='POSISI Y',                        t='CMM',           nom='99',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A050a', g='A050',   nm='POSISI X',                        t='CMM',           nom='7',      tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A050b', g='A050',   nm='POSISI Y',                        t='CMM',           nom='142',    tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B073a', g='B073',   nm='POSISI X',                        t='CMM',           nom='381',    tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B073b', g='B073',   nm='POSISI Y',                        t='CMM',           nom='8',      tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B051a', g='B051',   nm='POSISI X',                        t='CMM',           nom='50.5',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B051b', g='B051',   nm='POSISI Y',                        t='CMM',           nom='54.5',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B052a', g='B052',   nm='POSISI X',                        t='CMM',           nom='50.5',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B052b', g='B052',   nm='POSISI Y',                        t='CMM',           nom='54.5',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B053a', g='B053',   nm='POSISI X',                        t='CMM',           nom='22',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B053b', g='B053',   nm='POSISI Y',                        t='CMM',           nom='77',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B054a', g='B054',   nm='POSISI X',                        t='CMM',           nom='117',    tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B054b', g='B054',   nm='POSISI Y',                        t='CMM',           nom='83.5',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B055a', g='B055',   nm='POSISI X',                        t='CMM',           nom='258.5',  tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B055b', g='B055',   nm='POSISI Y',                        t='CMM',           nom='87.2',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B056a', g='B056',   nm='POSISI X',                        t='CMM',           nom='287.5',  tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B056b', g='B056',   nm='POSISI Y',                        t='CMM',           nom='86.7',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    # PAGE 4
    dict(id='B057a', g='B057',   nm='POSISI X',                        t='CMM',           nom='157.5',  tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B057b', g='B057',   nm='POSISI Y',                        t='CMM',           nom='83.5',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B058a', g='B058',   nm='POSISI X',                        t='CMM',           nom='62.5',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B058b', g='B058',   nm='POSISI Y',                        t='CMM',           nom='78',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B060a', g='B060',   nm='POSISI X',                        t='CMM',           nom='86',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B060b', g='B060',   nm='POSISI Y',                        t='CMM',           nom='86',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B061a', g='B061',   nm='POSISI X',                        t='CMM',           nom='146.5',  tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B061b', g='B061',   nm='POSISI Y',                        t='CMM',           nom='0',      tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B062a', g='B062',   nm='POSISI X',                        t='CMM',           nom='256',    tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B062b', g='B062',   nm='POSISI Y',                        t='CMM',           nom='94.5',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B063a', g='B063',   nm='POSISI X',                        t='CMM',           nom='372',    tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B063b', g='B063',   nm='POSISI Y',                        t='CMM',           nom='56.8',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B064a', g='B064',   nm='POSISI X',                        t='CMM',           nom='66.3',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B064b', g='B064',   nm='POSISI Y',                        t='CMM',           nom='44',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B065a', g='B065',   nm='POSISI X',                        t='CMM',           nom='75.1',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='B065b', g='B065',   nm='POSISI Y',                        t='CMM',           nom='26.2',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A040a', g='A040',   nm='POSISI X',                        t='CMM',           nom='63.5',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A040b', g='A040',   nm='POSISI Y',                        t='CMM',           nom='8.5',    tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A041a', g='A041',   nm='POSISI X',                        t='CMM',           nom='77.5',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A041b', g='A041',   nm='POSISI Y',                        t='CMM',           nom='55',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A042a', g='A042',   nm='POSISI X',                        t='CMM',           nom='30',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A042b', g='A042',   nm='POSISI Y',                        t='CMM',           nom='127',    tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A043a', g='A043',   nm='POSISI X',                        t='CMM',           nom='39',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A043b', g='A043',   nm='POSISI Y',                        t='CMM',           nom='118.5',  tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A044a', g='A044',   nm='POSISI X',                        t='CMM',           nom='68',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A044b', g='A044',   nm='POSISI Y',                        t='CMM',           nom='60',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A045a', g='A045',   nm='POSISI X',                        t='CMM',           nom='8',      tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A045b', g='A045',   nm='POSISI Y',                        t='CMM',           nom='70',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    # PAGE 5
    dict(id='A046a', g='A046',   nm='POSISI X',                        t='CMM',           nom='57.1',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A046b', g='A046',   nm='POSISI Y',                        t='CMM',           nom='56.6',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A047a', g='A047',   nm='POSISI X',                        t='CMM',           nom='6.5',    tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A047b', g='A047',   nm='POSISI Y',                        t='CMM',           nom='63',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A071a', g='A071',   nm='POSISI X',                        t='CMM',           nom='110',    tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A071b', g='A071',   nm='POSISI Y',                        t='CMM',           nom='84.5',   tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A151a', g='A151',   nm='POSISI X',                        t='CMM',           nom='36',     tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A151b', g='A151',   nm='POSISI Y',                        t='CMM',           nom='31',     tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A151c', g='A151',   nm='POSISI X TAP',                    t='CMM',           nom='36',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A151d', g='A151',   nm='POSISI Y TAP',                    t='CMM',           nom='31',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A152a', g='A152',   nm='POSISI X',                        t='CMM',           nom='38',     tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A152b', g='A152',   nm='POSISI Y',                        t='CMM',           nom='31',     tp=0.05,  tm=0.05,  kp=True,  mv=None),
    dict(id='A152c', g='A152',   nm='POSISI X TAP',                    t='CMM',           nom='38',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A152d', g='A152',   nm='POSISI Y TAP',                    t='CMM',           nom='31',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A111a', g='A111',   nm='KERATAAN / FLATNESS',             t='CMM',           nom='',       tp=None,  tm=None,  kp=False, mv=0.05),
    dict(id='A111b', g='A111',   nm='SUDUT FACE COMP',                 t='CRM',           nom="10°",    tp=None,  tm=None,  kp=False, mv=None),
    dict(id='A153a', g='A153',   nm='POSISI X',                        t='CMM',           nom='19',     tp=0.15,  tm=0.15,  kp=False, mv=None),
    dict(id='A153b', g='A153',   nm='POSISI Y',                        t='CMM',           nom='15',     tp=0.20,  tm=0.20,  kp=False, mv=None),
    dict(id='A153c', g='A153',   nm='HEIGHT SPOT FACE',                t='CMM',           nom='64',     tp=0.10,  tm=0.10,  kp=False, mv=None),
    dict(id='A153d', g='A153',   nm='ROUGHNESS SPOT FACE',             t='ROUGHNESS TEST',nom='MAX 25S',tp=None,  tm=None,  kp=False, mv=None),
    dict(id='A154a', g='A154',   nm='ANGLE',                           t='CMM',           nom="10°",    tp=None,  tm=None,  kp=False, mv=None),
    dict(id='A154b', g='A154',   nm='POSISI Y',                        t='CMM',           nom='15',     tp=0.20,  tm=0.20,  kp=False, mv=None),
]

HDR_DEF = dict(
    unitProduksi='MACHINING CRANK CASE',
    namaPart='',
    noPart='',
    line='',
    shift='2',
    tanggal='',
    noDies='',
    namaOperator='',
    nrp='',
    noDoc='',
    tglBerlaku='',
)

# ─────────────────────────────────────────────────────────────────────────────
# KONFIGURASI HEADER PER PART + MODEL
# Key = "{PartName}_{ModelName}" — harus sama persis dengan nilai di CSV.
# Field yang tidak dicantumkan akan jatuh ke HDR_DEF.
# ─────────────────────────────────────────────────────────────────────────────
HDR_CONFIG: dict[str, dict] = {
    'CRCS L_K1AL L1': dict(
        unitProduksi = 'MACHINING CRANK CASE',
        namaPart     = 'CRANK CASE COMP LEFT',
        noPart       = '11201-K1AL-N800-MA',
        line         = '1',
        noDoc        = '',
        tglBerlaku   = '',
    ),
    'CRCS R_K1AL L1': dict(
        unitProduksi = 'MACHINING CRANK CASE',
        namaPart     = 'CRANK CASE COMP RIGHT',
        noPart       = '11101-K1AL-N800-MA',
        line         = '1',
        noDoc        = '',
        tglBerlaku   = '',
    ),
    'CRCS L_K1AL L2': dict(
        unitProduksi = 'MACHINING CRANK CASE',
        namaPart     = 'CRANK CASE COMP LEFT',
        noPart       = '11201-K1AL-N800-MA',
        line         = '2',
        noDoc        = '',
        tglBerlaku   = '',
    ),
    'CRCS R_K1AL L2': dict(
        unitProduksi = 'MACHINING CRANK CASE',
        namaPart     = 'CRANK CASE COMP RIGHT',
        noPart       = '11101-K1AL-N800-MA',
        line         = '2',
        noDoc        = '',
        tglBerlaku   = '',
    ),
    'CRCS L_K1AL L3': dict(
        unitProduksi = 'MACHINING CRANK CASE',
        namaPart     = 'CRANK CASE COMP LEFT',
        noPart       = '11201-K1AL-N800-MA',
        line         = '3',
        noDoc        = '',
        tglBerlaku   = '',
    ),
    'CRCS R_K1AL L3': dict(
        unitProduksi = 'MACHINING CRANK CASE',
        namaPart     = 'CRANK CASE COMP RIGHT',
        noPart       = '11101-K1AL-N800-MA',
        line         = '3',
        noDoc        = '',
        tglBerlaku   = '',
    ),
    'CRCS L_K2SA': dict(
        unitProduksi = 'MACHINING CRANK CASE',
        namaPart     = 'CRANK CASE COMP LEFT',
        noPart       = '11200-K1N -N000',
        line         = '5 & 6',
        noDoc        = '64CK-0K2S-105-B00',
        tglBerlaku   = '01-Nov-21',
    ),
    'CRCS R_K2SA': dict(
        unitProduksi = 'MACHINING CRANK CASE',
        namaPart     = 'CRANK CASE COMP RIGHT',
        noPart       = '11100-K1N-N000',
        line         = '5 & 6',
        noDoc        = '64CK-0K2S-105-B00',
        tglBerlaku   = '01-Nov-21',
    ),
    'CRCS L_K60': dict(
        unitProduksi = 'MACHINING CRANK CASE',
        namaPart     = 'CRANK CASE COMP LEFT',
        noPart       = '11200-K60-B000',
        line         = '4',
        noDoc        = '64CK-OK6R-402-B01',
        tglBerlaku   = '01-Nov-17',
    ),
    'CRCS R_K60': dict(
        unitProduksi = 'MACHINING CRANK CASE',
        namaPart     = 'CRANK CASE COMP RIGHT',
        noPart       = '11100-K59-A1OO',
        line         = '4',
        noDoc        = '64CK-OK6R-507-B01',
        tglBerlaku   = '01-Nov-17',
    ),
    'MISSION CASE_K60': dict(
        unitProduksi = 'MACHINING CRANK CASE',
        namaPart     = 'CASE COMP MISSION',
        noPart       = '21200-KZR-6000',
        line         = '4',
        noDoc        = '71CK-OK6R-602-B01',
        tglBerlaku   = '01-Nov-17',
    ),
    # ── CYL COMP ──────────────────────────────────────────────────────────────
    'CYL COMP_K1AL': dict(
        unitProduksi = 'MACHINING CYLINDER COMP',
        namaPart     = 'CYLINDER COMP (K1AL)',
        noPart       = '12100-K1A -N800',
        line         = '1 & 2',
        noDoc        = '60MK-0K1A-100-B00',
        tglBerlaku   = '18-Mar-24',
    ),
    'CYL COMP_K60': dict(
        unitProduksi = 'MACHINING CYLINDER COMP K2VG',
        namaPart     = 'CYLINDER COMP (MA)',
        noPart       = '1210A-KZR -6006-IN',
        line         = '3',
        noDoc        = '',
        tglBerlaku   = '',
    ),
    'CYL COMP_K2SA': dict(
        unitProduksi = 'MACHINING CYLINDER COMP K2SA',
        namaPart     = 'CYLINDER COMP (MA)',
        noPart       = '1210A-K1Z -N000-IN',
        line         = '3',
        noDoc        = '60MK - 0K2S - 200 - B00',
        tglBerlaku   = '21-Aug-24',
    ),
    # ── HOLDER WATER PUMP ────────────────────────────────────────────────────
    'HOLDER WATER PUMP_K60': dict(
        unitProduksi = 'MACHINING CYLINDER HEAD',
        namaPart     = 'HOLDER COMP WATER PUMP',
        noPart       = '12211 - KWN - 9000',
        line         = '1',
        noDoc        = '62HK-0K6R-101-B01',
        tglBerlaku   = 'Dec-15',
    ),
    # ── CYL HEAD — K60 ───────────────────────────────────────────────────────
    'CYL HEAD GV_K60': dict(
        unitProduksi = 'MACHINING CYLINDER HEAD',
        namaPart     = 'HEAD ASSY, CYLINDER',
        noPart       = '1220B-K1A -N601-DL',
        line         = '2',
        noDoc        = '62HK-0KJA-104-B01',
        tglBerlaku   = 'Dec-23',
    ),
    'CYL HEAD CAM_K60': dict(
        unitProduksi = 'MACHINING CYLINDER HEAD',
        namaPart     = 'HEAD COMP,CYLINDER',
        noPart       = '1220C-K60R-B605-IN',
        line         = '2',
        noDoc        = '62HK-0K6R-101-B01',
        tglBerlaku   = 'Dec-15',
    ),
    'CYL HEAD NT_K60': dict(
        unitProduksi = 'MACHINING CYLINDER HEAD',
        namaPart     = 'HEAD COMP,CYLINDER',
        noPart       = '1220C-K60R-B605-IN',
        line         = '2',
        noDoc        = '62HK-0K6R-101-B01',
        tglBerlaku   = 'Dec-15',
    ),
    'CYL HEAD ROUGH_K60': dict(
        unitProduksi = 'MACHINING CYLINDER HEAD',
        namaPart     = 'CYLINDER HEAD ROUGH',
        noPart       = '',
        line         = '',
        noDoc        = '',
        tglBerlaku   = '',
    ),
    # ── CYL HEAD — K2SA ──────────────────────────────────────────────────────
    'CYL HEAD GV_K2SA': dict(
        unitProduksi = 'CYLINDER HEAD',
        namaPart     = 'HEAD COMP. CYLINDER',
        noPart       = '12200-K2SA-N004',
        line         = '1',
        noDoc        = '',
        tglBerlaku   = '',
    ),
    'CYL HEAD CAM_K2SA': dict(
        unitProduksi = 'CYLINDER HEAD',
        namaPart     = 'HEAD COMP. CYLINDER',
        noPart       = '12200-K2SA-N004',
        line         = '1',
        noDoc        = '',
        tglBerlaku   = '',
    ),
    'CYL HEAD NT_K2SA': dict(
        unitProduksi = 'CYLINDER HEAD',
        namaPart     = 'HEAD COMP. CYLINDER',
        noPart       = '1220B-K2S -N001-DL',
        line         = '1',
        noDoc        = '62HK-0K2S-104-B02',
        tglBerlaku   = 'Mar-23',
    ),
    'CYL HEAD ROUGH_K2SA': dict(
        unitProduksi = 'CYLINDER HEAD',
        namaPart     = 'HEAD COMP. CYLINDER',
        noPart       = '1220B-K2S -N000-DL',
        line         = '1',
        noDoc        = '62HK-0K2S-104-B02',
        tglBerlaku   = 'Jan-22',
    ),
    # ── CYL HEAD — K1AL ──────────────────────────────────────────────────────
    'CYL HEAD GV_K1AL': dict(
        unitProduksi = 'MACHINING CYLINDER HEAD',
        namaPart     = 'HEAD ASSY, CYLINDER',
        noPart       = '1220B-K1A -N601-DL',
        line         = '2 & 3',
        noDoc        = '62HK-0KJA-104-B01',
        tglBerlaku   = 'Dec-23',
    ),
    'CYL HEAD GV_K1AL L2': dict(
        unitProduksi = 'MACHINING CYLINDER HEAD',
        namaPart     = 'HEAD ASSY, CYLINDER',
        noPart       = '1220B-K1A -N601-DL',
        line         = '2',
        noDoc        = '62HK-0KJA-104-B01',
        tglBerlaku   = 'Dec-23',
    ),
    'CYL HEAD GV_K1AL L3': dict(
        unitProduksi = 'MACHINING CYLINDER HEAD',
        namaPart     = 'HEAD ASSY, CYLINDER',
        noPart       = '1220B-K1A -N601-DL',
        line         = '3',
        noDoc        = '62HK-0KJA-104-B01',
        tglBerlaku   = 'Dec-23',
    ),
    'CYL HEAD CAM_K1AL L2': dict(
        unitProduksi = 'MACHINING CYLINDER HEAD',
        namaPart     = 'HEAD ASSY, CYLINDER',
        noPart       = '1220B-K1A -N601-DL',
        line         = '2',
        noDoc        = '62HK-0KJA-104-B01',
        tglBerlaku   = 'Dec-23',
    ),
    'CYL HEAD NT_K1AL L2': dict(
        unitProduksi = 'MACHINING CYLINDER HEAD',
        namaPart     = 'HEAD ASSY, CYLINDER',
        noPart       = '1220B-K1A -N601-DL',
        line         = '2',
        noDoc        = '62HK-0KJA-104-B01',
        tglBerlaku   = 'Dec-23',
    ),
    'CYL HEAD CAM_K1AL L3': dict(
        unitProduksi = 'MACHINING CYLINDER HEAD',
        namaPart     = 'HEAD ASSY, CYLINDER',
        noPart       = '1220B-K1A -N601-DL',
        line         = '3',
        noDoc        = '62HK-0KJA-104-B01',
        tglBerlaku   = 'Dec-23',
    ),
    'CYL HEAD NT_K1AL L3': dict(
        unitProduksi = 'MACHINING CYLINDER HEAD',
        namaPart     = 'HEAD ASSY, CYLINDER',
        noPart       = '1220B-K1A -N601-DL',
        line         = '3',
        noDoc        = '62HK-0KJA-104-B01',
        tglBerlaku   = 'Dec-23',
    ),
    'CYL HEAD ROUGH_K1AL': dict(
        unitProduksi = 'MACHINING CYLINDER HEAD',
        namaPart     = 'HEAD ASSY, CYLINDER ( K0JF )',
        noPart       = '12200B-K0J -N002 - IN',
        line         = '2 & 3',
        noDoc        = '62HK-0KJA-104-B01',
        tglBerlaku   = 'Sep-19',
    ),
    'CYL HEAD ROUGH_K1AL L2': dict(
        unitProduksi = 'MACHINING CYLINDER HEAD',
        namaPart     = 'HEAD ASSY, CYLINDER ( K0JF )',
        noPart       = '12200B-K0J -N002 - IN',
        line         = '2',
        noDoc        = '62HK-0KJA-104-B01',
        tglBerlaku   = 'Sep-19',
    ),
    'CYL HEAD ROUGH_K1AL L3': dict(
        unitProduksi = 'MACHINING CYLINDER HEAD',
        namaPart     = 'HEAD ASSY, CYLINDER ( K0JF )',
        noPart       = '12200B-K0J -N002 - IN',
        line         = '3',
        noDoc        = '62HK-0KJA-104-B01',
        tglBerlaku   = 'Sep-19',
    ),
    # ── CYL COMP K1AL per line ──────────────────────────────────────────────
    'CYL COMP_K1AL L1': dict(
        unitProduksi = 'MACHINING CYLINDER COMP',
        namaPart     = 'CYLINDER COMP (K1AL)',
        noPart       = '12100-K1A -N800',
        line         = '1',
        noDoc        = '60MK-0K1A-100-B00',
        tglBerlaku   = '18-Mar-24',
    ),
    'CYL COMP_K1AL L2': dict(
        unitProduksi = 'MACHINING CYLINDER COMP',
        namaPart     = 'CYLINDER COMP (K1AL)',
        noPart       = '12100-K1A -N800',
        line         = '2',
        noDoc        = '60MK-0K1A-100-B00',
        tglBerlaku   = '18-Mar-24',
    ),
    'CYL COMP_K1AL L3': dict(
        unitProduksi = 'MACHINING CYLINDER COMP',
        namaPart     = 'CYLINDER COMP (K1AL)',
        noPart       = '12100-K1A -N800',
        line         = '3',
        noDoc        = '60MK-0K1A-100-B00',
        tglBerlaku   = '18-Mar-24',
    ),
    # Tambahkan kombinasi Part+Model lain di sini ↓
    # TODO: 'CRCS L_K2VJ', 'CRCS R_K2VJ', 'CYL COMP_K2VJ' — isi kalau data K2VJ ada
}

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def std_str(item):
    if item['mv'] is not None:
        return f"MAX {item['mv']}"
    if item['nom'] and item['tp'] is not None:
        return f"{item['nom']}  ±{item['tp']}"
    return item['nom'] or '—'

def is_oot(val, item):
    """
    Cek OOT berdasarkan nilai DEVIASI.
    - mv type (flatness): abs(deviation) > mv
    - nominal ± tol type: deviation di luar range [-tm, +tp]
    """
    try:
        v = float(val)
    except (TypeError, ValueError):
        return False
    if item.get('mv') is not None:
        return abs(v) > item['mv']
    tp = item.get('tp')
    tm = item.get('tm')
    if tp is not None and tm is not None:
        return not (-float(tm) <= v <= float(tp))
    return False

def init_meas():
    return {i['id']: {'n1': '', 'n2': '', 'n3': '', 'n4': '', 'n5': ''} for i in ITEMS}

def count_oot(report):
    """Jumlah nilai measurement yang NG (per cell, bukan per item)."""
    m      = report['measurements']
    items  = report.get('items', ITEMS)
    n_samp = len(report.get('samples', ['n1','n2','n3','n4','n5']))
    keys   = [f'n{i+1}' for i in range(n_samp)]
    return sum(
        1 for item in items
        for k in keys
        if is_oot(m.get(item['id'], {}).get(k, ''), item)
    )

def count_filled(report):
    """Jumlah titik inspeksi yang sudah ada minimal satu nilai terisi."""
    m      = report['measurements']
    items  = report.get('items', ITEMS)
    n_samp = len(report.get('samples', ['n1','n2','n3','n4','n5']))
    keys   = [f'n{i+1}' for i in range(n_samp)]
    return sum(
        1 for item in items
        if any(m.get(item['id'], {}).get(k, '') != '' for k in keys)
    )

def count_total_meas(report):
    """Total titik inspeksi (bukan items × samples)."""
    return len(report.get('items', ITEMS))

def count_kp_ng(report):
    """Jumlah nilai KP yang NG (per cell)."""
    m      = report['measurements']
    items  = report.get('items', ITEMS)
    n_samp = len(report.get('samples', ['n1','n2','n3','n4','n5']))
    keys   = [f'n{i+1}' for i in range(n_samp)]
    return sum(
        1 for item in items
        for k in keys
        if item.get('kp') and is_oot(m.get(item['id'], {}).get(k, ''), item)
    )

# Folder ilustrasi — sesuaikan dengan struktur folder proyekmu
# Konvensi nama file: {ModelName}.png, misal K60.png / K2VJ.png
from pathlib import Path as _Path
ILUSTRASI_DIR = str(_Path(__file__).resolve().parent.parent / "assets" / "ilustrasi")

def _load_ilustrasi_b64(model: str) -> str:
    """Return base64 string gambar ilustrasi untuk model, atau '' kalau tidak ada."""
    import base64
    if not model:
        return ''
    for ext in ['.png', '.jpg', '.jpeg', '.PNG', '.JPG']:
        path = _Path(ILUSTRASI_DIR) / f"{model}{ext}"
        if path.is_file():
            return base64.b64encode(path.read_bytes()).decode()
    return ''

def _load_ilustrasi(report: dict) -> None:
    """Patch report['ilustrasi_img'] jika belum ada."""
    if not report.get('ilustrasi_img'):
        h          = report.get('header', {})
        model      = h.get('modelName', '')
        part_name  = h.get('partName', h.get('namaPart', ''))
        report['ilustrasi_img'] = _get_ilustrasi_b64(model, part_name)

def today_str():
    d = datetime.now()
    return f"{d.day:02d}/{d.month:02d}/{d.year}"

# ─────────────────────────────────────────────────────────────────────────────
# DUMMY DATA
# ─────────────────────────────────────────────────────────────────────────────

def _dummy_meas():
    import random
    m = init_meas()
    for item in ITEMS:
        if item['tp'] is not None:
            nom = float(item['nom']) if item['nom'] and item['nom'].replace('.','').replace('-','').isdigit() else 0
            for n in ['n1','n2','n3','n4','n5']:
                val = nom + round(random.uniform(-item['tp']*0.9, item['tp']*0.9), 3)
                m[item['id']][n] = str(val)
        elif item['mv'] is not None:
            for n in ['n1','n2','n3','n4','n5']:
                m[item['id']][n] = str(round(random.uniform(0, item['mv']*0.85), 3))
    return m

DUMMY_REPORTS = [
    {
        'id': '1001',
        'createdAt': '2026-04-14T07:00:00',
        'submittedBy': 'Budi Santoso',
        'header': {**HDR_DEF, 'tanggal': '14/04/2026', 'shift': '1', 'namaOperator': 'Budi Santoso', 'nrp': '12345', 'noDies': 'D-001'},
        'measurements': _dummy_meas(),
    },
    {
        'id': '1002',
        'createdAt': '2026-04-14T15:00:00',
        'submittedBy': 'Siti Rahayu',
        'header': {**HDR_DEF, 'tanggal': '14/04/2026', 'shift': '2', 'namaOperator': 'Siti Rahayu', 'nrp': '67890', 'noDies': 'D-001'},
        'measurements': _dummy_meas(),
    },
]

# ─────────────────────────────────────────────────────────────────────────────
# CSV / DATABASE LOADER
# Ganti load_cmm_data() dengan query ke database Anda nanti
# ─────────────────────────────────────────────────────────────────────────────



@st.cache_data
def load_cmm_data(csv_path: str) -> pd.DataFrame:
    """Baca CSV CMM. Ganti dengan query DB nanti."""
    df = pd.read_csv(csv_path, dtype={'SampleNo': str})
    df['Date'] = pd.to_datetime(df['Date'])
    df['DateOnly'] = df['Date'].dt.date.astype(str)
    return df

def get_report_options(df: pd.DataFrame) -> pd.DataFrame:
    """Daftar laporan yang bisa dibuat dari data CSV (Category=Produksi)."""
    prod = df[df['Category'] == 'Produksi']
    opts = (prod.groupby(['DateOnly', 'Shift', 'PartName', 'ModelName'])
                .size().reset_index(name='rows'))
    return opts

def build_report_from_csv(df: pd.DataFrame, model: str, date_str: str, shift: int,
                          part_name: str = '') -> dict:
    """
    Buat report dict dari CSV untuk model+partname+tanggal+shift tertentu.
    Hasilnya sama persis dengan struktur report manual.
    """


    # Filter data — sertakan PartName kalau diberikan
    mask = (
        (df['ModelName'] == model) &
        (df['DateOnly']  == date_str) &
        (df['Shift']     == shift) &
        (df['Category']  == 'Produksi')
    )
    if part_name:
        mask = mask & (df['PartName'] == part_name)
    data = df[mask].copy()

    if data.empty:
        return None

    # Ambil info header dari baris pertama
    first = data.iloc[0]
    tanggal_dt = pd.to_datetime(date_str)
    tanggal_fmt = f"{tanggal_dt.day:02d}/{tanggal_dt.month:02d}/{tanggal_dt.year}"
    _part_name = part_name or str(first['PartName'])

    # Derive samples dari data aktual — urutan kemunculan pertama
    samples = list(dict.fromkeys(
        data.sort_values(['Date', 'Cycle'])['SampleNo'].astype(str).tolist()
    ))
    if not samples:
        samples = ['N1']

    # Lookup HDR_CONFIG dengan key "{PartName}_{ModelName}"
    _cfg_key = f"{_part_name}_{model}"
    _part_cfg = HDR_CONFIG.get(_cfg_key, {})
    header = {
        **HDR_DEF,
        **_part_cfg,                       # override per part+model dari HDR_CONFIG
        'tanggal':   tanggal_fmt,
        'shift':     str(shift),
        'partName':  str(first['PartName']),  # PartName mentah dari CSV (untuk lookup)
        'namaPart':  _part_cfg.get('namaPart', _part_name),  # prefer HDR_CONFIG, fallback CSV
        'modelName': model,
    }

    # Pivot Actual
    pivot = (data.pivot_table(
        index=['ref', 'ID', 'point', 'Nominal', 'Uppertol', 'Lowertol', 'KP'],
        columns='SampleNo',
        values='Actual',
        aggfunc='first'
    ).reset_index())
    for s in samples:
        if s not in pivot.columns:
            pivot[s] = ''

    # Pivot Deviation (untuk template Excel)
    pivot_dev = (data.pivot_table(
        index=['ref', 'ID', 'point'],
        columns='SampleNo',
        values='Deviation',
        aggfunc='first'
    ).reset_index())
    for s in samples:
        if s not in pivot_dev.columns:
            pivot_dev[s] = ''
    # Build dict: iid → {n1:dev, n2:dev, ...}
    dev_meas_tmp = {}
    for _, row in pivot_dev.iterrows():
        iid = str(row['ID']).strip()
        meas = {}
        for ni, s in enumerate(samples):
            v = row.get(s, '')
            meas[f'n{ni+1}'] = (
                '' if pd.isna(v) or v == ''
                else f"{float(v):.4f}".rstrip('0').rstrip('.')
            )
        dev_meas_tmp[iid] = meas

    # Build items + measurements dari pivot Actual
    items = []
    measurements       = {}
    deviation_measurements = {}
    for _, row in pivot.iterrows():
        ref   = str(row['ref']).strip()
        iid   = str(row['ID']).strip()
        point = str(row['point']).strip()
        nom   = row['Nominal']
        tp    = row['Uppertol']
        tm    = row['Lowertol']
        kp    = bool(row['KP'])

        if nom == 0 and tm == 0:
            mv = tp; tp_val = None; tm_val = None
        else:
            mv = None; tp_val = tp; tm_val = abs(tm)

        item = dict(
            id=iid, g=ref if ref != '-' else iid,
            nm=point, t='CMM', nom=str(nom),
            tp=tp_val, tm=tm_val, kp=kp, mv=mv,
        )
        items.append(item)

        meas = {}
        for ni, s in enumerate(samples):
            v = row.get(s, '')
            meas[f'n{ni+1}'] = (
                '' if pd.isna(v) or v == ''
                else f"{float(v):.4f}".rstrip('0').rstrip('.')
            )
        measurements[iid]           = dev_meas_tmp.get(iid, meas)
        deviation_measurements[iid] = dev_meas_tmp.get(iid, meas)  # alias, same

    # Cari ilustrasi berdasarkan part+model (bukan model saja)
    _ilustrasi_img  = _get_ilustrasi_b64(model, _part_name)
    _ilustrasi_path = ''
    # Cari file path dari _ILUSTRASI_MAP pakai key yang sama dengan _get_ilustrasi_b64
    _ikey = f"{model}_{_part_name}".replace(' ', '_').upper()
    _best_len, _best_file = 0, ''
    for _mk, _fname in _ILUSTRASI_MAP.items():
        if _ikey.startswith(_mk.upper()) and len(_mk) > _best_len:
            _best_len, _best_file = len(_mk), _fname
    if _best_file:
        _p = _ASSETS_DIR / _best_file
        if _p.exists():
            _ilustrasi_path = str(_p)

    return {
        'id':                    f"{model}_{_part_name}_{date_str}_{shift}".replace(' ', '_'),
        'createdAt':             pd.Timestamp.now().isoformat(),
        'submittedBy':           'CMM Auto',
        'header':                header,
        'items':                 items,
        'samples':               samples,
        'sample_order':          list(samples),
        'measurements':          measurements,
        'deviation_measurements': deviation_measurements,
        'ilustrasi_img':         _ilustrasi_img,
        'ilustrasi_path':        _ilustrasi_path,
    }





# Generates the exact visual replica of the AHM QCL report
# ─────────────────────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────────────────────
# ILUSTRASI — mapping model+part → file di assets/ilustrasi/
# Key: "{MODEL}_{PART}" dengan spasi diganti _ (sama persis dengan descriptive.py)
# ─────────────────────────────────────────────────────────────────────────────

_ASSETS_DIR = _Path(__file__).resolve().parent.parent / "assets" / "ilustrasi"

# ── Sama persis dengan _IMG_MAP di descriptive.py ─────────────────────────
_ILUSTRASI_MAP: dict[str, str] = {
    "K2VJ_CYL_COMP":    "K2VJ.png",
    "K60_CYL_COMP":     "K2V.png",
    "K60_CRCS_L":       "K60.png",
    "K2SA_CYL_COMP":    "K2SA_CYLCOMP.jpg",
    "K1AL_L1_CRCS_L":   "K1AL_L1.jpg",
    "K1AL_L1_CRCS_R":   "K1AL_L1_R.jpg",
    "K1AL_L2_CRCS_L":   "K1AL_L2.jpg",
    "K1AL_L2_CRCS_R":   "K1AL_L2_R.jpg",
    "K1AL_L3_CRCS_L":   "K1AL_L3.jpg",
    "K1AL_L3_CRCS_R":   "K1AL_L3_R.jpg",
    "K2SA_CRCS_L":      "K2SA_CRCS_L.jpg",
    "K2SA_CRCS_R":      "K2SA_CRCS_R.jpg",
    "K60_CRCS_R":       "K60_R.jpg",
    "K60_MISSION":      "K60_MISSION.jpg",
    "K1AL_CYL_COMP":    "K1AL_CYLCOMP.jpg",
    "K60_GV":           "K60_GV.jpg",
    "K2SA_GV":          "K2SA_GV.jpg",
    "K1AL_GV":          "K1AL_GV.jpg",
    "K60_CAM":          "K60_CAM.jpg",
    "K2SA_CAM":         "K2SA_CAM.jpg",
    "K1AL_L2_CAM":      "K1AL_CAM.jpg",
    "K1AL_L3_CAM":      "K1AL_CAM.jpg",
    "K60_NT":           "K60_NT.jpg",
    "K2SA_NT":          "K2SA_NT.jpg",
    "K1AL_L2_NT":       "K1AL_NT.jpg",
    "K1AL_L3_NT":       "K1AL_NT.jpg",
    "K60_ROUGH":        "K60_ROUGH.jpg",
    "K2SA_ROUGH":       "K2SA_ROUGH.jpg",
    "K1AL_ROUGH":       "K1AL_ROUGH.jpg",
    "K60_HWP":          "K60_WP.jpg",
}

_ILUSTRASI_B64_CACHE: dict[str, str] = {}

def _detect_ilustrasi_key(part: str, model: str) -> str:
    """
    Deteksi active_key untuk ilustrasi — IDENTIK dengan _detect_active_key
    di descriptive.py agar file yang diload selalu sama.
    """
    p, m = part.lower(), model.lower()
    # CRCS
    if "k2vj" in m and "cyl comp" in p:       return "K2VJ_CYL_COMP"
    if "k60"  in m and "crcs l"   in p:       return "K60_CRCS_L"
    if "k60"  in m and "crcs r"   in p:       return "K60_CRCS_R"
    if "k60"  in m and "mission"  in p:       return "K60_MISSION"
    if "k2sa" in m and "crcs l"   in p:       return "K2SA_CRCS_L"
    if "k2sa" in m and "crcs r"   in p:       return "K2SA_CRCS_R"
    if "k1al l1" in m and "crcs l" in p:      return "K1AL_L1_CRCS_L"
    if "k1al l1" in m and "crcs r" in p:      return "K1AL_L1_CRCS_R"
    if "k1al l2" in m and "crcs l" in p:      return "K1AL_L2_CRCS_L"
    if "k1al l2" in m and "crcs r" in p:      return "K1AL_L2_CRCS_R"
    if "k1al l3" in m and "crcs l" in p:      return "K1AL_L3_CRCS_L"
    if "k1al l3" in m and "crcs r" in p:      return "K1AL_L3_CRCS_R"
    # CYL COMP
    if "k2sa" in m and "cyl comp" in p:       return "K2SA_CYL_COMP"
    if "k1al" in m and "cyl comp" in p:       return "K1AL_CYL_COMP"
    if "k60"  in m and "cyl comp" in p:       return "K60_CYL_COMP"
    # CYL HEAD GV
    if "k60"  in m and "gv" in p:            return "K60_GV"
    if "k2sa" in m and "gv" in p:            return "K2SA_GV"
    if "k1al" in m and "gv" in p:            return "K1AL_GV"
    # CYL HEAD CAM
    if "k60"     in m and "cam" in p:         return "K60_CAM"
    if "k2sa"    in m and "cam" in p:         return "K2SA_CAM"
    if "k1al l2" in m and "cam" in p:         return "K1AL_L2_CAM"
    if "k1al l3" in m and "cam" in p:         return "K1AL_L3_CAM"
    # CYL HEAD NT
    if "k60"     in m and "nt" in p:          return "K60_NT"
    if "k2sa"    in m and "nt" in p:          return "K2SA_NT"
    if "k1al l2" in m and "nt" in p:          return "K1AL_L2_NT"
    if "k1al l3" in m and "nt" in p:          return "K1AL_L3_NT"
    # CYL HEAD ROUGH
    if "k60"  in m and "rough" in p:          return "K60_ROUGH"
    if "k2sa" in m and "rough" in p:          return "K2SA_ROUGH"
    if "k1al" in m and "rough" in p:          return "K1AL_ROUGH"
    # HOLDER WATER PUMP
    if "k60"  in m and "water pump" in p:     return "K60_HWP"
    return ""

def _get_ilustrasi_b64(model: str, part_name: str) -> str:
    """Cari ilustrasi — pakai key system yang sama dengan descriptive.py."""
    active_key = _detect_ilustrasi_key(part_name, model)
    if not active_key:
        return ''

    fname = _ILUSTRASI_MAP.get(active_key, '')
    if not fname:
        return ''

    if fname in _ILUSTRASI_B64_CACHE:
        return _ILUSTRASI_B64_CACHE[fname]

    img_path = _ASSETS_DIR / fname
    if not img_path.exists():
        return ''

    import base64
    ext    = img_path.suffix.lower()
    mime   = 'image/png' if ext == '.png' else 'image/jpeg'
    result = f'data:{mime};base64,{base64.b64encode(img_path.read_bytes()).decode()}'
    _ILUSTRASI_B64_CACHE[fname] = result
    return result


def build_report_html(report) -> str:
    """Render laporan QCL sebagai HTML read-only (view mode)."""
    h = report['header']
    m = report['measurements']
    illustration = report.get('illustration', '')  # base64 image or empty

    # Pre-load confidential image DULU biar bisa dipakai di stamp area
    import base64 as _b64_rpt
    from pathlib import Path as _rpt_path
    _conf_img_html = ""
    for _conf_fname in ["confidential.png", "confidential.jpg", "CONFIDENTIAL.png", "CONFIDENTIAL.jpg"]:
        _conf_path = _rpt_path("assets") / _conf_fname
        if _conf_path.exists():
            _conf_ext  = _conf_path.suffix.lower()
            _conf_mime = "image/png" if _conf_ext == ".png" else "image/jpeg"
            _conf_b64  = _b64_rpt.b64encode(_conf_path.read_bytes()).decode()
            _conf_img_html = (
                f'<img src="data:{_conf_mime};base64,{_conf_b64}" '
                f'style="max-height:72px;max-width:100%;object-fit:contain;display:block;margin:0 auto;" />'
            )
            break

    # Build illustration HTML — tampil di area stamp (logo/stamp) di header
    if illustration:
        illustration_html = f'<img src="data:image/png;base64,{illustration}" style="max-width:100%; max-height:80px; object-fit:contain;" />'
    elif _conf_img_html:
        illustration_html = _conf_img_html   # ← gambar confidential masuk ke stamp area
    else:
        illustration_html = '<div style="font-size:8px; color:#bbb; margin-top:20px;">(logo/stamp)</div>'

    # Build ILUSTRASI section HTML — auto-load dari assets/ilustrasi
    _model    = h.get('modelName', '')
    _part     = h.get('partName',  h.get('namaPart', ''))
    _img_data = report.get('ilustrasi_img', '') or _get_ilustrasi_b64(_model, _part)
    if _img_data:
        # strip prefix kalau sudah ada (dari cache _get_ilustrasi_b64 sudah include prefix)
        if _img_data.startswith('data:'):
            _img_src = _img_data
        else:
            _img_src = f'data:image/jpeg;base64,{_img_data}'
        ilustrasi_section = f'<img src="{_img_src}" style="max-width:100%; max-height:260px; object-fit:contain;" />'
    else:
        ilustrasi_section = '<div style="font-size:11px; color:#ccc; padding-top:100px;">[ Ilustrasi tidak tersedia ]</div>'

    # Use dynamic items from CSV report, or fall back to static ITEMS
    items_to_render  = report.get('items', ITEMS)
    samples_data     = report.get('samples', ['N1','N2','N3','N4','N5'])  # data keys (n1..nN)
    sample_order     = report.get('sample_order', samples_data)           # display labels (reorderable)
    # Guard: sample_order bisa beda panjang dgn samples_data (rename/edit)
    if len(sample_order) != len(samples_data):
        sample_order = list(samples_data)  # fallback ke samples asli
    n_samples        = len(sample_order)
    samples_to_render = sample_order  # used for TH labels
    sample_labels    = {f'n{i+1}': sample_order[i] for i in range(n_samples)}

    # Compute group rowspans
    span = {}
    for item in items_to_render:
        span[item['g']] = span.get(item['g'], 0) + 1

    # Build inspection rows
    seen = {}
    rows_html = ''
    for item in items_to_render:
        is_first = item['g'] not in seen
        if is_first:
            seen[item['g']] = True

        meas = m.get(item['id'], {})
        cell_style_base = 'padding:2px 5px; border:1px solid #888; font-size:9.5px; vertical-align:middle;'

        # Sample value cells (read-only)
        sample_cells = ''
        for n in [f'n{i+1}' for i in range(len(samples_to_render))]:
            val = meas.get(n, '')
            oot = is_oot(val, item)
            val_style = cell_style_base + 'text-align:center; font-family:Arial,monospace; background:#f0f4ff;'
            if oot:
                if item.get('kp'):
                    val_style += 'color:#92400E; font-weight:bold; background:#FEF9C3;'  # KP: kuning
                else:
                    val_style += 'color:#cc0000; font-weight:bold; background:#ffe8e8;'  # NG biasa: merah
            sample_cells += f'<td style="{val_style}">{val}</td>'

        # Group cell (rowspan on first)
        grp_cell = ''
        if is_first:
            grp_cell = f'''<td rowspan="{span[item['g']]}" style="
                {cell_style_base}
                text-align:center; font-weight:bold; font-family:Arial,monospace;
                background:#dce8ff; color:#0a1e45; border-right:2px solid #1a4080;
                font-size:9px; white-space:nowrap;">
                {item['g']}
            </td>'''

        nm_style = cell_style_base
        if item['kp']:
            nm_style += 'color:#92400E; font-weight:bold; background:#FFFBEB;'

        kp_cell = '<span style="background:#FEF08A;color:#713F12;font-weight:bold;font-size:8px;padding:1px 3px;border-radius:2px;">KP</span>' if item['kp'] else ''

        rows_html += f'''<tr>
            {grp_cell}
            <td style="{nm_style}">{item['nm']}</td>
            <td style="{cell_style_base} text-align:center; font-size:9px; color:#555;">{item['t']}</td>
            <td style="{cell_style_base} text-align:center; font-family:Arial,monospace; font-size:9px; white-space:nowrap;">{std_str(item)}</td>
            <td style="{cell_style_base} text-align:center;">{kp_cell}</td>
            <td style="{cell_style_base} text-align:center; font-size:9px; color:#555;">HB</td>
            {sample_cells}
        </tr>'''

    form_start = ''
    form_end   = ''
    save_btn   = ''

    # JS drag-and-drop: swap kolom dan tulis order baru ke URL query param
    _so_json = json.dumps(sample_order)
    rid      = report.get('id', 'rpt').replace('-','_')
    js = f'''
    <style>
      th.sample-th {{ cursor:grab; user-select:none; }}
      th.sample-th.drag-over {{ outline:2px solid #60a5fa !important; outline-offset:-2px; }}
    </style>
    <script>
    (function(){{
      var order = {_so_json};
      var src = null;
      document.querySelectorAll('th.sample-th').forEach(function(th){{
        th.draggable = true;
        th.addEventListener('dragstart', function(){{
          src = th.textContent.trim(); th.style.opacity='0.45';
        }});
        th.addEventListener('dragend', function(){{
          th.style.opacity='';
          document.querySelectorAll('th.sample-th').forEach(function(t){{t.classList.remove('drag-over');}});
        }});
        th.addEventListener('dragover', function(e){{
          e.preventDefault(); th.classList.add('drag-over');
        }});
        th.addEventListener('dragleave', function(){{
          th.classList.remove('drag-over');
        }});
        th.addEventListener('drop', function(e){{
          e.preventDefault(); th.classList.remove('drag-over');
          var tgt = th.textContent.trim();
          if (!src || src===tgt) return;
          // Swap DOM
          document.querySelectorAll('th.sample-th').forEach(function(t){{
            var txt=t.textContent.trim();
            if(txt===src) t.textContent=tgt;
            else if(txt===tgt) t.textContent=src;
          }});
          // Swap array
          var si=order.indexOf(src), ti=order.indexOf(tgt);
          if(si>=0&&ti>=0){{var tmp=order[si];order[si]=order[ti];order[ti]=tmp;}}
          src = null;
          // Tulis ke URL — satu key, value = rid:json agar tidak menumpuk
          try{{
            var u=new URL(window.parent.location.href);
            u.searchParams.set('_so', '{rid}:' + JSON.stringify(order));
            window.parent.history.replaceState({{}}, '', u.toString());
          }}catch(err){{}}
        }});
      }});
    }})();
    </script>
    '''

    # Pre-compute logo AHM sebagai base64
    _logo_rpt = ""
    _logo_svg = _rpt_path("assets/Logo_AHM.svg")
    if _logo_svg.exists():
        _logo_rpt = "data:image/svg+xml;base64," + _b64_rpt.b64encode(_logo_svg.read_bytes()).decode()
    _logo_cell = (f'<img src="{_logo_rpt}" style="max-height:44px;max-width:100%;width:auto;object-fit:contain;display:block;margin:0 auto;" />'
                  if _logo_rpt else
                  '<div style="color:#cc0000;font-weight:900;font-size:28px;letter-spacing:2px;">AHM</div>'
                  '<div style="font-size:10px;color:#333;font-weight:700;">PT Astra Honda Motor</div>')

    html = f'''
    <!DOCTYPE html>
    <html>
    <head>
    <meta charset="UTF-8">
    <style>
        * {{ box-sizing: border-box; margin:0; padding:0; }}
        body {{ font-family: Arial, sans-serif; font-size:10px; background:white; }}
        table {{ border-collapse: collapse; }}
        input:focus {{ border-color:#1a4080 !important; box-shadow:0 0 0 2px rgba(26,64,128,.2); }}
        @media print {{
            .no-print {{ display:none !important; }}
            body {{ background:white; -webkit-print-color-adjust:exact; print-color-adjust:exact; }}
            img {{ max-width:100% !important; }}
            thead {{ display:table-header-group; }}
            tr {{ page-break-inside:avoid; }}
        }}
    </style>
    </head>
    <body style="padding:12px 12px 4px 12px;">

    {form_start}

    <table style="width:100%; border-collapse:collapse; border:1.5px solid #333; table-layout:fixed;">
      <colgroup>
        <col style="width:10%;"> <col style="width:8%;">  <col style="width:10%;"> <col style="width:15%;"> <col style="width:10%;"> <col style="width:7%;">  <col style="width:10%;"> <col style="width:10%;"> <col style="width:10%;"> <col style="width:10%;"> </colgroup>

      <tr>
        <td colspan="3" rowspan="2" style="text-align:center; padding:10px 4px; border:1px solid #333; vertical-align:middle;">
          {_logo_cell}
        </td>
        <td colspan="4" rowspan="3" style="text-align:center; font-weight:900; font-size:16px; letter-spacing:0.5px; padding:7px 4px; border:1px solid #333; vertical-align:middle;">
          WORK STATION INSPECTION RESULT DATA QCL
        </td>
        <td colspan="2" style="padding:2px 5px; border:1px solid #333; font-size:9px; text-align:center; font-weight:bold;">NO.<br>DOKUMEN</td>
        <td colspan="1" style="padding:2px 5px; border:1px solid #333; font-weight:bold; font-size:10px; text-align:center;">{h.get('noDoc','')}</td>
      </tr>

      <tr>
        <td colspan="2" style="padding:2px 5px; border:1px solid #333; font-size:9px; text-align:center; font-weight:bold;">TGL<br>BERLAKU</td>
        <td colspan="1" style="padding:2px 5px; border:1px solid #333; font-size:10px; text-align:center; font-weight:bold;">{h.get('tglBerlaku','')}</td>
      </tr>

      <tr>
        <td colspan="3" style="padding:4px 6px; border:1px solid #333; font-size:9px; font-weight:bold;">PROCESS ENGINEERING B DEPT.</td>
        <td colspan="1" style="padding:4px 5px; border:1px solid #333; font-size:9px; text-align:center; font-weight:bold;">REV &nbsp; 0</td>
        <td colspan="1" style="padding:4px 5px; border:1px solid #333; font-size:9px; text-align:center; font-weight:bold;">HAL</td>
        <td colspan="1" style="padding:4px 5px; border:1px solid #333; font-size:9px; text-align:center; font-weight:bold;">1 DARI 5</td>
      </tr>

      <tr>
        <td colspan="2" style="padding:3px 5px; border:1px solid #333; font-size:9px; font-weight:bold; white-space:nowrap;">UNIT PRODUKSI</td>
        <td colspan="2" style="padding:3px 5px; border:1px solid #333; font-size:10px; font-weight:bold;">: {h.get('unitProduksi','')}</td>
        <td colspan="2" style="padding:3px 5px; border:1px solid #333; font-size:11px; font-weight:900; background:#FFFF00; color:#000; text-align:center; letter-spacing:1px;">CONFIDENTIAL STATUS</td>
        <td colspan="1" style="border:1px solid #333;"></td>
        <td colspan="1" style="padding:3px 4px; border:1px solid #333; font-size:10px; font-weight:bold; text-align:center;">QCL 1</td>
        <td colspan="1" style="padding:3px 4px; border:1px solid #333; font-size:10px; font-weight:bold; text-align:center;">QCL 2</td>
        <td colspan="1" style="padding:3px 4px; border:1px solid #333; font-size:10px; font-weight:bold; text-align:center;">QCL 3</td>
      </tr>

      <tr>
        <td colspan="2" style="padding:3px 5px; border:1px solid #333; font-size:9px; font-weight:bold; white-space:nowrap;">NAMA PART</td>
        <td colspan="2" style="padding:3px 5px; border:1px solid #333; font-size:10px; font-weight:bold;">: {h.get('namaPart','')}</td>
        <td colspan="2" rowspan="5" style="border:1px solid #333; padding:6px; vertical-align:middle; text-align:center;">
          {illustration_html}
        </td>
        <td colspan="1" style="padding:3px 5px; border:1px solid #333; font-size:9px; font-weight:bold; text-align:center;">TANGGAL</td>
        <td colspan="1" style="border:1px solid #333;"></td>
        <td colspan="1" style="border:1px solid #333;"></td>
        <td colspan="1" style="border:1px solid #333;"></td>
      </tr>

      <tr>
        <td colspan="2" style="padding:3px 5px; border:1px solid #333; font-size:9px; font-weight:bold; white-space:nowrap;">NO. PART</td>
        <td colspan="2" style="padding:3px 5px; border:1px solid #333; font-size:10px; font-weight:bold;">: {h.get('noPart','')}</td>
        <td colspan="1" rowspan="3" style="padding:3px 5px; border:1px solid #333; font-size:9px; font-weight:bold; text-align:center; vertical-align:middle;">TANDA<br>TANGAN</td>
        <td colspan="1" rowspan="3" style="border:1px solid #333;"></td>
        <td colspan="1" rowspan="3" style="border:1px solid #333;"></td>
        <td colspan="1" rowspan="3" style="border:1px solid #333;"></td>
      </tr>

      <tr>
        <td colspan="2" style="padding:3px 5px; border:1px solid #333; font-size:9px; font-weight:bold; white-space:nowrap;">LINE</td>
        <td colspan="2" style="padding:3px 5px; border:1px solid #333; font-size:10px; font-weight:bold;">: {h.get('line','')}</td>
      </tr>

      <tr>
        <td colspan="2" style="padding:3px 5px; border:1px solid #333; font-size:9px; font-weight:bold; white-space:nowrap;">SHIFT</td>
        <td colspan="2" style="padding:3px 5px; border:1px solid #333; font-size:10px; font-weight:bold;">: {h.get('shift','')}</td>
      </tr>

      <tr>
        <td colspan="2" style="padding:3px 5px; border:1px solid #333; font-size:9px; font-weight:bold; white-space:nowrap;">TANGGAL</td>
        <td colspan="2" style="padding:3px 5px; border:1px solid #333; font-size:10px; font-weight:bold;">
            <div style="display:flex; justify-content:space-between;">
                <span>: {h.get('tanggal','')}</span>
                <span style="font-size:9px;">NO DIES: {h.get('noDies','')}</span>
            </div>
        </td>
        <td colspan="1" style="padding:3px 5px; border:1px solid #333; font-size:9px; font-weight:bold; text-align:center;">NAMA /<br>NRP</td>
        <td colspan="1" style="border:1px solid #333; font-size:9px; text-align:center; font-weight:bold; vertical-align:bottom; padding-bottom:2px;">
            {h.get('namaOperator','')}<br>{h.get('nrp','')}
        </td>
        <td colspan="1" style="border:1px solid #333;"></td>
        <td colspan="1" style="border:1px solid #333;"></td>
      </tr>

    </table>

    <!-- ══ ILUSTRASI ════════════════════════════════════════════════════ -->
    <div style="border:1px solid #333; border-top:none; padding:6px 8px;
         text-align:center; background:#fafafa; min-height:50px;">
      {ilustrasi_section}
    </div>

    <!-- ══ INSPECTION TABLE ═══════════════════════════════════════════ -->
    <table style="width:100%; border-collapse:collapse; margin-top:0;">
      <thead>
        <tr>
          <th style="background:#0a1e45; color:#cce0ff; padding:5px 6px; border:1px solid #333; font-size:9px; text-align:center; width:65px;">NO</th>
          <th style="background:#0a1e45; color:#cce0ff; padding:5px 6px; border:1px solid #333; font-size:9px; text-align:left; min-width:175px;">POINT PEMERIKSAAN</th>
          <th style="background:#0a1e45; color:#cce0ff; padding:5px 6px; border:1px solid #333; font-size:9px; text-align:center; width:80px;">INSPECTION TOOL</th>
          <th style="background:#0a1e45; color:#cce0ff; padding:5px 6px; border:1px solid #333; font-size:9px; text-align:center; width:115px;">STANDARD (mm)</th>
          <th style="background:#0a1e45; color:#cce0ff; padding:5px 4px; border:1px solid #333; font-size:9px; text-align:center; width:70px;">KRITIKAL POINT</th>
          <th style="background:#0a1e45; color:#cce0ff; padding:5px 4px; border:1px solid #333; font-size:9px; text-align:center; width:70px;">KRITERIA SAFETY</th>
          {
            "".join(
                f'<th class="sample-th" style="background:#122850;color:#cce0ff;padding:5px 4px;border:1px solid #333;font-size:9px;text-align:center;width:52px;">{s}</th>'
                for s in sample_order
            )
          }
        </tr>
      </thead>
      <tbody>
        {rows_html}
      </tbody>
    </table>

    {save_btn}
    {form_end}
    {js}
    </body>
    </html>
    '''
    return html

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT (openpyxl)
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(report) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'QCL Report'

    h = report['header']
    m = report['measurements']

    # Styles
    def bd(style='thin'):
        s = Side(style=style)
        return Border(left=s, right=s, top=s, bottom=s)

    def fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def font(bold=False, size=9, color='000000'):
        return Font(bold=bold, size=size, color=color, name='Arial')

    def align(h='center', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    navy  = fill('0A1E45')
    navy2 = fill('122850')
    red   = fill('A32D2D')
    yellow= fill('FFFF00')
    blue_light = fill('DCE8FF')
    sample_bg  = fill('F0F4FF')
    oot_bg     = fill('FFE8E8')
    white = fill('FFFFFF')

    # ── HEADER ROWS ────────────────────────────────────────────────────────
    row = 1

    # Row 1: AHM logo area + title + doc number
    ws.merge_cells(f'A{row}:A{row+2}')    # AHM logo
    ws.merge_cells(f'B{row}:G{row}')      # Title
    ws.merge_cells(f'H{row}:H{row}')
    ws.merge_cells(f'I{row}:I{row}')

    c = ws.cell(row, 1, 'AHM'); c.font = Font(bold=True, size=14, color='FFFFFF', name='Arial'); c.fill = red; c.alignment = align()
    c = ws.cell(row, 2, 'WORK STATION INSPECTION RESULT DATA QCL')
    c.font = font(bold=True, size=12); c.alignment = align(); c.border = bd()
    ws.cell(row, 8, 'NO. DOKUMEN').font = font(size=8, color='555555')
    ws.cell(row, 9, h.get('noDoc','')).font = font(bold=True)
    for col in range(1, 10):
        ws.cell(row, col).border = bd()
    row += 1

    # Row 2: Dept + tgl berlaku
    ws.merge_cells(f'B{row}:G{row}')
    c = ws.cell(row, 2, 'PROCESS ENGINEERING B DEPT.'); c.font = Font(italic=True, size=8, color='555555', name='Arial'); c.border = bd()
    ws.cell(row, 8, 'TGL BERLAKU').font = font(size=8, color='555555')
    ws.cell(row, 9, h.get('tglBerlaku','')).font = font()
    for col in [1,8,9]: ws.cell(row, col).border = bd()
    row += 1

    # Row 3: CONFIDENTIAL
    ws.merge_cells(f'B{row}:G{row}')
    c = ws.cell(row, 2, 'CONFIDENTIAL STATUS'); c.font = font(bold=True, size=10); c.fill = yellow; c.border = bd(); c.alignment = align('left')
    ws.cell(row, 8, 'REV').font = font(size=8, color='555555')
    ws.cell(row, 9, '0').font = font()
    for col in [1,8,9]: ws.cell(row, col).border = bd()
    row += 1

    # Row 4: Unit Produksi + QCL sigs + Tanggal
    ws.merge_cells(f'B{row}:D{row}')
    ws.cell(row, 1, 'UNIT PRODUKSI').font = font(size=8, color='555555')
    ws.cell(row, 2, h.get('unitProduksi','')).font = font(bold=True)
    for col in [5,6,7]:
        ws.cell(row, col, f'QCL {col-4}').font = font(size=8, color='555555')
        ws.cell(row, col).alignment = align()
    ws.cell(row, 8, 'TANGGAL').font = font(size=8, color='555555')
    ws.cell(row, 9, h.get('tanggal','')).font = font(bold=True)
    for col in range(1,10): ws.cell(row, col).border = bd()
    row += 1

    # Rows 5-7: Nama Part / No Part / Line + signatures + Nama/NRP / No Dies / Shift
    ws.merge_cells(f'E{row}:G{row+2}')  # signature area
    sig_cell = ws.cell(row, 5, 'Tanda Tangan')
    sig_cell.font = font(size=8, color='555555'); sig_cell.alignment = align()

    ws.cell(row, 1, 'NAMA PART').font = font(size=8, color='555555')
    ws.merge_cells(f'B{row}:D{row}')
    ws.cell(row, 2, h.get('namaPart','')).font = font(bold=True)
    ws.cell(row, 8, 'NAMA / NRP').font = font(size=8, color='555555')
    ws.cell(row, 9, f"{h.get('namaOperator','')} / {h.get('nrp','')}").font = font(bold=True)
    for col in [1,2,8,9]: ws.cell(row, col).border = bd()
    row += 1

    ws.cell(row, 1, 'NO. PART').font = font(size=8, color='555555')
    ws.merge_cells(f'B{row}:D{row}')
    ws.cell(row, 2, h.get('noPart','')).font = Font(bold=True, size=11, name='Courier New')
    ws.cell(row, 8, 'NO DIES').font = font(size=8, color='555555')
    ws.cell(row, 9, h.get('noDies','')).font = font(bold=True)
    for col in [1,2,8,9]: ws.cell(row, col).border = bd()
    row += 1

    ws.cell(row, 1, 'LINE').font = font(size=8, color='555555')
    ws.cell(row, 2, h.get('line','')).font = font(bold=True)
    ws.cell(row, 3, 'SHIFT').font = font(size=8, color='555555')
    ws.cell(row, 4, h.get('shift','')).font = Font(bold=True, size=18, color='0A1E45', name='Arial')
    ws.cell(row, 8, 'FREK. PEMERIKSAAN').font = font(size=8, color='555555')
    ws.cell(row, 9, '1 / SHIFT').font = font(bold=True)
    for col in [1,2,3,4,5,8,9]: ws.cell(row, col).border = bd()
    row += 1

    # ── TABLE HEADER ───────────────────────────────────────────────────────
    header_row = row
    _report_samples = report.get('sample_order', report.get('samples', ['N1', 'N2', 'N3', 'N4', 'N5']))
    cols = ['NO', 'POINT PEMERIKSAAN', 'INSPECTION TOOL', 'STANDARD (mm)', 'KRITIKAL POINT', 'KRITERIA SAFETY'] + list(_report_samples)
    for ci, label in enumerate(cols, 1):
        c = ws.cell(row, ci, label)
        c.font = Font(bold=True, size=9, color='CCE0FF', name='Arial')
        c.fill = navy2 if ci >= 7 else navy
        c.alignment = align()
        c.border = bd()
    row += 1

    # ── INSPECTION DATA ROWS ──────────────────────────────────────────────
    _items_to_use = report.get('items', ITEMS)
    span  = {}
    for item in _items_to_use:
        span[item['g']] = span.get(item['g'], 0) + 1

    grp_start = {}  # group -> first excel row

    for item in _items_to_use:
        meas = m.get(item['id'], {})
        is_first = item['g'] not in grp_start
        if is_first:
            grp_start[item['g']] = row

        # Group col (A) — filled only on first row, merged later
        if is_first:
            c = ws.cell(row, 1, item['g'])
            c.font = Font(bold=True, size=9, name='Courier New', color='0A1E45')
            c.fill = blue_light
            c.alignment = align()
            c.border = bd()

        # Point name
        c = ws.cell(row, 2, item['nm'])
        c.font = Font(bold=item['kp'], size=9, color='CC0000' if item['kp'] else '000000', name='Arial')
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = bd()

        # Tools
        c = ws.cell(row, 3, item['t']); c.font = font(size=8, color='555555'); c.alignment = align(); c.border = bd()

        # Standard
        c = ws.cell(row, 4, std_str(item)); c.font = Font(size=9, name='Courier New'); c.alignment = align(); c.border = bd()

        # KP
        c = ws.cell(row, 5, 'KP' if item['kp'] else '')
        c.font = Font(bold=True, size=8, color='CC0000', name='Arial'); c.alignment = align(); c.border = bd()

        # HB
        c = ws.cell(row, 6, 'HB'); c.font = font(size=8, color='555555'); c.alignment = align(); c.border = bd()

        # Sample columns (dynamic)
        for ni, sname in enumerate(_report_samples, 7):
            val_str = meas.get(f'n{ni-6}', '')
            oot = is_oot(val_str, item)
            try:
                val = float(val_str) if val_str else None
            except ValueError:
                val = val_str
            c = ws.cell(row, ni, val)
            c.fill = oot_bg if oot else sample_bg
            c.font = Font(bold=oot, size=9, color='CC0000' if oot else '000000', name='Courier New')
            c.alignment = align()
            c.border = bd()
            c.number_format = '0.000'

        row += 1

    # Merge group cells
    for grp, start_row in grp_start.items():
        if span[grp] > 1:
            ws.merge_cells(f'A{start_row}:A{start_row + span[grp] - 1}')
            c = ws.cell(start_row, 1)
            c.alignment = align()

    # Column widths (dynamic sample columns)
    col_widths = [9, 36, 14, 16, 5, 5] + [9] * len(_report_samples)
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row heights
    for r in range(1, row):
        ws.row_dimensions[r].height = 14
    ws.row_dimensions[header_row].height = 18

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Template path config ──────────────────────────────────────────
_TEMPLATE_DIR = _Path(__file__).resolve().parent.parent / "assets" / "templates"

EXCEL_TEMPLATES = {
    "K2VJ": str(_TEMPLATE_DIR / "K2VJ.xlsx"),
    "K60":  str(_TEMPLATE_DIR / "K60.xlsx"),
}

# Sample column mapping sesuai template K2VJ (row 34: Q=AL...Z=ER)
_K2VJ_SAMPLE_COLS = {
    'AL':17,'AR':18,'BL':19,'BR':20,'CL':21,
    'CR':22,'DL':23,'DR':24,'EL':25,'ER':26
}


def _write_cell(ws, row: int, col: int, value, number_format: str = None,
                fill=None, alignment=None):
    """
    Tulis ke cell di worksheet, handle merged cell otomatis.
    Kalau cell adalah bagian dari merge range, cari top-left dan tulis ke sana.
    """
    from openpyxl.cell import MergedCell as _MC
    cell = ws.cell(row, col)
    if isinstance(cell, _MC):
        # Cari top-left dari merge range ini
        for rng in ws.merged_cells.ranges:
            if (rng.min_row <= row <= rng.max_row and
                    rng.min_col <= col <= rng.max_col):
                cell = ws.cell(rng.min_row, rng.min_col)
                break
        else:
            return  # tidak ketemu, skip
    cell.value = value
    if number_format:
        cell.number_format = number_format
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment


def build_excel_from_template(report) -> bytes | None:
    """
    Export laporan ke Excel menggunakan template yang sudah ada.
    Isi nilai di baris ( ANGKA ) sesuai ref + parameter dari template.
    Return bytes, atau None kalau template tidak ada.
    """
    from pathlib import Path
    from datetime import datetime as _dt

    model    = report.get('header', {}).get('modelName', '')
    tpl_path = EXCEL_TEMPLATES.get(model)
    if not tpl_path or not Path(tpl_path).exists():
        return None

    try:
        wb = load_workbook(tpl_path)
    except Exception:
        return None

    ws      = wb.active
    m       = report.get('deviation_measurements') or report.get('measurements', {})
    items   = report.get('items', [])
    samples = report.get('sample_order', report.get('samples', []))

    # ── Sample → kolom template ───────────────────────────────────
    # Baca dari row 34 template (baris header sample)
    sample_col_map = {}
    for c in range(1, wb.active.max_column + 1):
        v = wb.active.cell(34, c).value
        if v and str(v).strip() in ['AL','AR','BL','BR','CL','CR','DL','DR','EL','ER']:
            sample_col_map[str(v).strip()] = c
    if not sample_col_map:
        sample_col_map = _K2VJ_SAMPLE_COLS  # fallback hardcode

    # ── Fill header ───────────────────────────────────────────────
    hdr = report.get('header', {})
    _write_cell(ws, 5, 6,  hdr.get('unitProduksi', ''))
    _write_cell(ws, 6, 6,  hdr.get('namaPart', ''))
    _write_cell(ws, 7, 6,  hdr.get('noPart', ''))
    _write_cell(ws, 8, 6,  hdr.get('line', ''))

    # Tanggal ke S11 (col 19)
    try:
        tgl = hdr.get('tanggal', '')
        if tgl:
            dt = _dt.strptime(tgl, '%d/%m/%Y')
            _write_cell(ws, 11, 19, dt, number_format='DD/MM/YYYY')
    except Exception:
        _write_cell(ws, 11, 19, hdr.get('tanggal', ''))

    # Shift ke O11 (col 15)
    try:
        shift_val = hdr.get('shift', '')
        curr = str(ws.cell(11, 15).value or '')
        new_val = f'SHIFT  :  {shift_val}' if 'SHIFT' in curr.upper() else shift_val
        _write_cell(ws, 11, 15, new_val)
    except Exception:
        pass

    # Nama / NRP ke X11 (col 24)
    _write_cell(ws, 11, 24, f"{hdr.get('namaOperator','')} / {hdr.get('nrp','')}")

    # ── Sisipkan ilustrasi ────────────────────────────────────────
    img_path = report.get('ilustrasi_path', '')
    if img_path:
        try:
            from openpyxl.drawing.image import Image as _XLImg
            xl_img = _XLImg(img_path)
            xl_img.anchor = 'B12'
            ws.add_image(xl_img)
        except Exception:
            pass

    # ── Build ref+param → row number dari template ────────────────
    # Cari semua baris '( ANGKA )' dan trace balik ref + param
    ref_param_to_row: dict = {}
    current_ref  = None
    last_param   = None

    for r in range(1, wb.active.max_row + 1):
        b_val = str(ws.cell(r, 2).value or '').strip()
        c_val = str(ws.cell(r, 3).value or '').strip()
        g_val = str(ws.cell(r, 7).value or '').strip()

        if b_val:
            current_ref = b_val.upper()
        if c_val:
            last_param = c_val.upper()

        if g_val == '( ANGKA )' and current_ref and last_param:
            ref_param_to_row[(current_ref, last_param)] = r

    # ── Isi nilai measurement (deviation) ────────────────────────
    # Tidak ada coloring — pertahankan format template asli
    for item in items:
        item_ref   = str(item.get('g', item.get('id',''))).strip().upper()
        item_param = str(item.get('nm', '')).strip().upper()
        item_meas  = m.get(item['id'], {})

        row_num = ref_param_to_row.get((item_ref, item_param))
        if not row_num:
            for (ref_k, param_k), rn in ref_param_to_row.items():
                if ref_k == item_ref and param_k.startswith(item_param[:10]):
                    row_num = rn
                    break
        if not row_num:
            continue

        for ni, sname in enumerate(samples):
            col = sample_col_map.get(sname)
            if not col:
                continue
            val_str = item_meas.get(f'n{ni+1}', '')
            if not val_str:
                continue
            try:
                val = float(val_str)
            except (ValueError, TypeError):
                continue
            _write_cell(ws, row_num, col, val,
                        number_format='0.0000')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_pdf_from_excel(report) -> bytes | None:
    """
    Convert laporan ke PDF via Microsoft Excel (win32com).
    Hanya berjalan di Windows dengan Excel installed.
    Return bytes PDF, atau None kalau gagal.
    """
    if not HAS_WIN32 or not HAS_OPENPYXL:
        return None

    import os, tempfile

    xl_bytes = build_excel_from_template(report) or build_excel(report)

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp.write(xl_bytes)
        xl_path = tmp.name

    pdf_path = xl_path.replace('.xlsx', '.pdf')

    try:
        pythoncom.CoInitialize()
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible        = False
        excel.DisplayAlerts  = False
        excel.ScreenUpdating = False
        wb = excel.Workbooks.Open(
            os.path.abspath(xl_path),
            UpdateLinks=False,
            ReadOnly=True
        )
        wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        wb.Close(False)
        excel.Quit()
        del excel
        with open(pdf_path, 'rb') as f:
            return f.read()
    except Exception as e:
        st.error(f"Gagal export PDF: {e}")
        return None
    finally:
        try: pythoncom.CoUninitialize()
        except: pass
        try: os.unlink(xl_path)
        except: pass
        try: os.unlink(pdf_path)
        except: pass




# ─────────────────────────────────────────────────────────────────────────────
# MAIN PAGE CLASS
# ─────────────────────────────────────────────────────────────────────────────

class ReportPage:
    def __init__(self, df_all: pd.DataFrame, current_user: str = 'Operator'):
        # DateOnly dihitung sekali di sini, bukan di dalam render method
        if not df_all.empty and 'DateOnly' not in df_all.columns:
            df_all = df_all.copy()
            df_all['DateOnly'] = pd.to_datetime(df_all['Date']).dt.date.astype(str)
        self.df_all       = df_all
        self.current_user = current_user
        self._init_state()

    def _init_state(self):
        defaults = {
            'report_page':    'wsird',
            'qcl_reports':    [],
            'qcl_view':       'list',
            'qcl_current':    None,
            'qcl_saved_data': None,
            'rpt_dirty':      True,    # flag cache laporan (ikut pola diagnostic.py)
        }
        for k, v in defaults.items():
            if k not in st.session_state:
                st.session_state[k] = v

    # ── Cache helpers (pola diagnostic.py) ────────────────────────────────

    def _get_reports_cached(self) -> list:
        """Ambil laporan dari DB, cache di session_state, invalidate via rpt_dirty."""
        role     = st.session_state.get('role', 'Operator')
        username = st.session_state.get('username', '')
        if st.session_state.get('rpt_dirty', True) or 'rpt_cached' not in st.session_state:
            st.session_state['rpt_cached'] = get_reports(role, username)
            st.session_state['rpt_dirty']  = False
        return st.session_state['rpt_cached']

    @staticmethod
    def _invalidate_rpt_cache():
        st.session_state['rpt_dirty'] = True

    
    def _get_html_cached(self, report: dict) -> str:
        """Cache HTML laporan di session_state — regenerasi hanya jika header berubah."""
        rid  = report['id']
        hkey = f'_rpt_html_{rid}'
        vkey = f'_rpt_html_v_{rid}'
        # versi = hash header (measurements tidak berubah di view-only flow)
        ver  = hash(json.dumps({'h': report.get('header', {}), 'so': report.get('sample_order', [])}, sort_keys=True, default=str))
        if st.session_state.get(vkey) != ver or hkey not in st.session_state:
            st.session_state[hkey] = build_report_html(report)
            st.session_state[vkey] = ver
        return st.session_state[hkey]

    
    def _get_excel_cached(self, report: dict) -> bytes:
        """Cache Excel bytes — regenerasi hanya jika header berubah."""
        rid  = report['id']
        hkey = f'_rpt_xl_{rid}'
        vkey = f'_rpt_xl_v_{rid}'
        ver  = hash(json.dumps({'h': report.get('header', {}), 'so': report.get('sample_order', [])}, sort_keys=True, default=str))
        if st.session_state.get(vkey) != ver or hkey not in st.session_state:
            st.session_state[hkey] = build_excel_from_template(report) or build_excel(report)
            st.session_state[vkey] = ver
        return st.session_state[hkey]

    def render(self):
        page = st.session_state.report_page
        if page in ('landing', 'wsird'):
            self._render_wsird()

    def _render_wsird(self):
        view = st.session_state.qcl_view
        if view == 'list':
            self._render_list()
        elif view == 'edit':
            self._render_edit()
        elif view == 'view':
            self._render_view()

    # ── LIST VIEW ──────────────────────────────────────────────────────────
    @st.fragment
    def _render_list(self):
        role     = st.session_state.get("role", "Operator")
        username = st.session_state.get("username", "")

        st.markdown(
            '<div class="page-hdr"><span class="page-title">Report</span>'
            '<span class="page-sub">WSIRD Produksi</span></div>'
            '<div class="section-desc">Buat laporan inspeksi QCL dari data CMM · kirim ke Produksi · export ke Excel atau PDF.</div>',
            unsafe_allow_html=True
        )
        st.divider()

        # Laporan dari cache (bukan query DB tiap rerun)
        all_reports = self._get_reports_cached()
        if not all_reports:
            # Coba load dari session state
            all_reports = st.session_state.get('qcl_reports', [])

        # ── Generate (hanya Measurement & Admin) ─────────────────────────
        if role != "Produksi":
          with st.expander("📂 Generate Laporan", expanded=True):
            if not self.df_all.empty:
                opts = get_report_options(self.df_all)
                if not opts.empty:
                    st.caption(f"{len(opts)} kombinasi tersedia")
                    opts['label']  = opts['PartName'] + ' · ' + opts['ModelName']
                    label_to_model = dict(zip(opts['label'], opts['ModelName']))
                    label_to_part  = dict(zip(opts['label'], opts['PartName']))

                    c1, c2, c3, c4 = st.columns(4)
                    sel_label  = c1.selectbox("Part / Model", opts['label'].unique().tolist(), key="sel_model")
                    sel_model  = label_to_model[sel_label]
                    sel_part   = label_to_part[sel_label]
                    date_opts  = sorted(
                        opts[(opts['ModelName'] == sel_model) &
                             (opts['PartName']  == sel_part)]['DateOnly'].unique().tolist(),
                        reverse=True
                    )
                    # key pakai sel_label agar selectbox reset saat model berubah
                    sel_date   = c2.selectbox("Tanggal", date_opts,
                                              key=f"sel_date_{sel_label}")
                    shift_opts = sorted(
                        opts[(opts['ModelName'] == sel_model) &
                             (opts['PartName']  == sel_part) &
                             (opts['DateOnly']  == sel_date)]['Shift']
                        .unique().tolist()
                    )
                    sel_shift  = c3.selectbox("Shift", shift_opts,
                                              format_func=lambda x: f"Shift {x}",
                                              key=f"sel_shift_{sel_label}_{sel_date}")

                    # Cek apakah laporan sudah ada (untuk info note saja)
                    candidate_id  = f"{sel_model}_{sel_part}_{sel_date}_{sel_shift}".replace(' ', '_')
                    existing_meta = get_report_meta(candidate_id)
                    if existing_meta:
                        _created = existing_meta.get('createdAt', '')[:10] if existing_meta else ''
                        st.caption(f"ℹ️ Sudah ada laporan untuk kombinasi ini"
                                   + (f" (dibuat {_created})" if _created else "")
                                   + " — klik Buat Laporan untuk timpa.")

                    if st.button("Buat Laporan", type="primary",
                                 use_container_width=True):
                            with st.spinner("Membuat laporan..."):
                                r = build_report_from_csv(
                                    self.df_all, sel_model, sel_date, int(sel_shift),
                                    part_name=sel_part)
                            if r:
                                with st.spinner("Menyimpan..."):
                                    rid = save_report(r, username)
                                r['id'] = rid
                                self._invalidate_rpt_cache()
                                reports  = st.session_state.qcl_reports
                                existing = next((i for i, x in enumerate(reports)
                                                 if x['id'] == rid), None)
                                if existing is not None:
                                    reports[existing] = r
                                else:
                                    reports.insert(0, r)
                                st.session_state.qcl_reports = reports
                                st.session_state.qcl_current = copy.deepcopy(r)
                                st.session_state.qcl_view    = 'edit'
                                st.rerun()
                            else:
                                st.warning("Tidak ada data untuk kombinasi yang dipilih.")
                else:
                    st.info("Tidak ada data Produksi tersedia.")
            else:
                st.info("Data CMM belum tersedia.")

        st.divider()

        # ── Load laporan dari DB ke session state ─────────────────────────
        db_reports = self._get_reports_cached()
        session_ids = {r['id'] for r in st.session_state.qcl_reports}
        for meta in db_reports:
            if meta['id'] not in session_ids:
                r_data = load_report_data(meta['id'])
                if r_data:
                    st.session_state.qcl_reports.append(r_data)

        all_reports = st.session_state.qcl_reports

        # Role Produksi hanya lihat laporan yang sudah Terkirim
        if role == "Produksi":
            all_reports = [r for r in all_reports
                           if (get_report_meta(r['id']) or {}).get('status','draft') == 'sent']

        # ── Filter + Sort + Pagination ────────────────────────────────────
        # Baris 1: Part·Model, Status, Shift
        fa1, fa2, fa3 = st.columns([3, 2, 2])
        with fa1:
            if not self.df_all.empty:
                combos = (self.df_all[["PartName","ModelName"]].dropna()
                          .drop_duplicates().sort_values(["PartName","ModelName"]))
                pm_opts = ["Semua Part & Model"] + [
                    f"{r.PartName} · {r.ModelName}" for _, r in combos.iterrows()
                ]
            else:
                pm_opts = ["Semua Part & Model"]
            flt_pm = st.selectbox("Part · Model", pm_opts, key="rpt_flt_pm") or "Semua Part & Model"
        with fa2:
            flt_status = st.selectbox(
                "Status", ["Semua Status","Draft","Terkirim"],
                key="rpt_flt_status"
            )
        with fa3:
            flt_shift = st.selectbox(
                "Shift", ["Semua Shift","Shift 1","Shift 2","Shift 3"],
                key="rpt_flt_shift"
            )

        # Baris 2: Tanggal, Urutkan
        fb1, fb2 = st.columns([2, 2])
        with fb1:
            _all_dates = sorted(
                {r['header'].get('tanggal','') for r in all_reports
                 if r['header'].get('tanggal','')},
                reverse=True
            )
            date_opts_flt = ["Semua Tanggal"] + _all_dates
            flt_date = st.selectbox("Tanggal", date_opts_flt, key="rpt_flt_date")
        with fb2:
            flt_sort = st.selectbox(
                "Urutkan",
                ["Terbaru", "Terlama", "Part A-Z", "NG Terbanyak"],
                key="rpt_flt_sort"
            )
        PER_PAGE = 25

        # Terapkan filter
        filtered = list(all_reports)
        if flt_pm != "Semua Part & Model":
            filtered = [r for r in filtered
                if f"{r['header'].get('partName', r['header'].get('namaPart',''))} · {r['header'].get('modelName','')}" == flt_pm]
        if flt_shift != "Semua Shift":
            filtered = [r for r in filtered
                if str(r['header'].get('shift','')) == flt_shift.replace('Shift ', '')]
        if flt_date != "Semua Tanggal":
            filtered = [r for r in filtered
                if r['header'].get('tanggal','') == flt_date]
        if flt_status != "Semua Status":
            s_map = {"Terkirim":"sent","Draft":"draft"}
            target = s_map.get(flt_status,"")
            def _get_status(r):
                m = get_report_meta(r['id'])
                return m.get("status","draft") if m else "draft"
            filtered = [r for r in filtered if _get_status(r) == target]

        # Terapkan sort
        def _parse_tgl(r):
            t = r['header'].get('tanggal','')
            try:
                from datetime import datetime as _dt
                return _dt.strptime(t, '%d/%m/%Y')
            except Exception:
                return _dt.min
        if flt_sort == "Terbaru":
            filtered.sort(key=_parse_tgl, reverse=True)
        elif flt_sort == "Terlama":
            filtered.sort(key=_parse_tgl, reverse=False)
        elif flt_sort == "Part A-Z":
            filtered.sort(key=lambda r: (
                r['header'].get('partName', r['header'].get('namaPart','')),
                r['header'].get('modelName','')
            ))
        elif flt_sort == "NG Terbanyak":
            filtered.sort(key=lambda r: count_oot(r), reverse=True)

        # Pagination
        total       = len(filtered)
        total_pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)

        # Reset halaman kalau filter berubah
        _flt_sig = f"{flt_pm}|{flt_status}|{flt_shift}|{flt_date}|{flt_sort}"
        if st.session_state.get('_rpt_flt_sig') != _flt_sig:
            st.session_state['_rpt_flt_sig'] = _flt_sig
            st.session_state['rpt_page'] = 1

        cur_page = st.session_state.get('rpt_page', 1)
        cur_page = max(1, min(cur_page, total_pages))

        # Header: count + nav
        hc1, hc2, hc3, hc4, hc5 = st.columns([2, 1, 2, 1, 2])
        hc1.caption(f"{total} laporan · halaman {cur_page}/{total_pages}")
        if hc2.button("<", key="rpt_prev", disabled=(cur_page <= 1)):
            st.session_state['rpt_page'] = cur_page - 1
            st.rerun()
        hc3.markdown(
            f"<div style='text-align:center;font-size:12px;padding-top:6px;color:#64748B;'>"
            f"Hal. {cur_page} / {total_pages}</div>",
            unsafe_allow_html=True
        )
        if hc4.button(">", key="rpt_next", disabled=(cur_page >= total_pages)):
            st.session_state['rpt_page'] = cur_page + 1
            st.rerun()
        hc5.empty()

        page_reports = filtered[(cur_page-1)*PER_PAGE : cur_page*PER_PAGE]

        for r in page_reports:
            ng       = count_oot(r)
            kp_ng    = count_kp_ng(r)
            filled   = count_filled(r)
            _n_total = count_total_meas(r)

            meta_db   = get_report_meta(r['id'])
            db_status = meta_db.get("status", "draft") if meta_db else "draft"
            _pname = r['header'].get('partName') or r['header'].get('namaPart', '')
            _mname = r['header'].get('modelName', '')
            nm = f"{_pname} {_mname}".strip() or r['header'].get('noPart') or r['id']

            # Badge NG/OK (inline kiri)
            if ng > 0:
                ng_badge = '<span style="background:#FEE2E2;color:#991B1B;font-size:10px;font-weight:600;padding:2px 8px;border-radius:10px;">NG</span>'
            elif filled > 0:
                ng_badge = '<span style="background:#DCFCE7;color:#166534;font-size:10px;font-weight:600;padding:2px 8px;border-radius:10px;">OK</span>'
            else:
                ng_badge = ''

            # Badge status workflow (kanan)
            if db_status == "sent":
                status_badge = '<span style="background:#DBEAFE;color:#1E40AF;font-size:10px;font-weight:600;padding:2px 10px;border-radius:10px;">📤 Terkirim</span>'
            else:
                status_badge = '<span style="background:#F1F5F9;color:#64748B;font-size:10px;font-weight:600;padding:2px 10px;border-radius:10px;">📝 Draft</span>'

            # info NG inline
            ng_info = ''
            if ng > 0:
                ng_info += f' &nbsp;·&nbsp; <span style="color:#DC2626;font-weight:600;">🔴 {ng} NG</span>'
            if kp_ng > 0:
                ng_info += f' &nbsp;·&nbsp; <span style="color:#D97706;font-weight:600;">⚠ {kp_ng} KP NG</span>'

            st.markdown(f"""
            <div style="background:white;border:1px solid #E2E8F0;border-radius:10px;
                 padding:14px 20px;margin-bottom:4px;
                 box-shadow:0 1px 3px rgba(15,23,42,.05);">
              <div style="display:flex;align-items:center;justify-content:space-between;">
                <div style="display:flex;align-items:center;gap:8px;font-size:15px;font-weight:700;color:#0F172A;">
                  {nm}
                  <span style="font-weight:500;color:#64748B;font-size:13px;">· Shift {r['header']['shift']}</span>
                  {ng_badge}
                </div>
                <div>{status_badge}</div>
              </div>
              <div style="font-size:11px;color:#64748B;margin-top:4px;">
                📅 {r['header']['tanggal']} &nbsp;·&nbsp;
                👤 {r['header'].get('namaOperator','—')} &nbsp;·&nbsp;
                {filled}/{_n_total} titik{ng_info}
              </div>
            </div>
            """, unsafe_allow_html=True)

            c1, c2, c3, c4 = st.columns([1, 1, 1, 0.4])
            with c1:
                if st.button("Lihat", key=f"view_{r['id']}", use_container_width=True):
                    st.session_state.qcl_current = copy.deepcopy(r)
                    st.session_state.qcl_view    = 'view'
                    st.rerun()
            with c2:
                if role != "Produksi":
                    if st.button("Edit", key=f"edit_{r['id']}", use_container_width=True):
                        st.session_state.qcl_current = copy.deepcopy(r)
                        st.session_state.qcl_view    = 'edit'
                        st.rerun()
            with c3:
                if st.button("PDF", key=f"pdf_{r['id']}", use_container_width=True):
                    st.session_state.qcl_current = copy.deepcopy(r)
                    st.session_state['_print_rpt'] = r['id']
                    st.session_state.qcl_view = 'view'
                    st.rerun()
            with c4:
                if st.button("Hapus", key=f"del_{r['id']}", use_container_width=True,
                             help="Hapus laporan"):
                    rid = r['id']
                    st.session_state.qcl_reports = [
                        x for x in st.session_state.qcl_reports if x['id'] != rid
                    ]
                    from local_db import delete_report as _del_rpt
                    _del_rpt(rid)
                    self._invalidate_rpt_cache()
                    st.rerun()
            st.markdown('<div style="height:8px;"></div>', unsafe_allow_html=True)

        # Pagination bawah
        if total_pages > 1:
            bc1, bc2, bc3, bc4, bc5 = st.columns([2, 1, 2, 1, 2])
            if bc2.button("Prev", key="rpt_prev_bot", disabled=(cur_page <= 1)):
                st.session_state['rpt_page'] = cur_page - 1; st.rerun()
            bc3.markdown(
                f"<div style='text-align:center;font-size:12px;padding-top:6px;color:#64748B;'>"
                f"Hal. {cur_page} / {total_pages}</div>", unsafe_allow_html=True
            )
            if bc4.button("Next", key="rpt_next_bot", disabled=(cur_page >= total_pages)):
                st.session_state['rpt_page'] = cur_page + 1; st.rerun()

    # ── EDIT VIEW (header only) ────────────────────────────────────────────
    @st.fragment
    def _render_edit(self):
        r = st.session_state.qcl_current
        if not r:
            st.session_state.qcl_view = 'list'; st.rerun()

        components.html("""<script>
setTimeout(function(){
  try{
    var d=window.parent.document;
    ['[data-testid="stMainBlockContainer"]','[data-testid="stMain"]',
     '.main','section.main','.block-container'].forEach(function(s){
      var e=d.querySelector(s); if(e) e.scrollTop=0;
    });
    window.parent.scrollTo(0,0);
    d.documentElement.scrollTop=0; d.body.scrollTop=0;
  }catch(e){}
},120);
</script>""", height=0)
        # Status badge untuk indikator draft
        _meta_edit  = get_report_meta(r['id'])
        _db_status  = _meta_edit.get('status', 'draft') if _meta_edit else 'draft'
        _status_cfg = {
            'draft': ('📝 Draft',   '#F1F5F9', '#475569'),
            'sent':  ('📤 Terkirim','#DBEAFE', '#1E40AF'),
        }
        _slabel, _sbg, _sclr = _status_cfg.get(_db_status, _status_cfg['draft'])

        col_back, col_title, col_save = st.columns([1, 5, 1.5])
        with col_back:
            if st.button("Kembali"):
                st.session_state.qcl_view = 'list'; st.rerun()
        with col_title:
            _part_lbl = (r['header'].get('partName') or r['header'].get('namaPart',''))
            _model_lbl = r['header'].get('modelName','')
            _shift_lbl = r['header'].get('shift','')
            st.markdown(
                f"**Edit** · {_part_lbl} {_model_lbl} · Shift {_shift_lbl} "
                f'&nbsp;<span style="background:{_sbg};color:{_sclr};'
                f'font-size:11px;font-weight:600;padding:2px 10px;'
                f'border-radius:10px;vertical-align:middle;">{_slabel}</span>',
                unsafe_allow_html=True
            )
        with col_save:
            if st.button("Simpan", type="primary", use_container_width=True):
                # Baca _so: format "rid:json_order" — cek rid cocok dulu
                _so_raw = st.query_params.get('_so', '')
                if _so_raw and ':' in _so_raw:
                    try:
                        _so_rid, _so_json = _so_raw.split(':', 1)
                        if _so_rid == r['id'].replace('-', '_'):
                            _new_so = json.loads(_so_json)
                            if isinstance(_new_so, list) and _new_so:
                                r['sample_order'] = _new_so
                    except Exception:
                        pass
                with st.spinner("Menyimpan..."):
                    self._save_current()
                # Auto kirim ke Produksi + notif NG
                _uname = st.session_state.get("username", "")
                update_report_status(r['id'], 'sent')
                n_notif = _send_ng_notifs_from_report(r, _uname)
                # Invalidate cache
                rid = r['id']
                for pfx in ('_rpt_html_', '_rpt_html_v_', '_rpt_xl_', '_rpt_xl_v_'):
                    st.session_state.pop(f'{pfx}{rid}', None)
                self._invalidate_rpt_cache()
                _msg = "Tersimpan & terkirim ke Produksi ✓"
                if n_notif > 0:
                    _msg += f" · {n_notif} notifikasi NG"
                st.success(_msg)
                st.session_state.qcl_view = 'view'
                st.rerun()

        st.divider()

        h = r['header']
        # Baris 1: info part (config header)
        c1, c2, c3, c4 = st.columns(4)
        h['unitProduksi'] = c1.text_input("Unit Produksi", h.get('unitProduksi', ''))
        h['namaPart']     = c2.text_input("Nama Part",     h.get('namaPart', ''))
        h['noPart']       = c3.text_input("No. Part",      h.get('noPart', ''))
        h['line']         = c4.text_input("Line",          h.get('line', ''))
        # Baris 2: info operator
        c1, c2, c3, c4 = st.columns(4)
        h['namaOperator'] = c1.text_input("Nama Operator", h.get('namaOperator', ''))
        h['nrp']          = c2.text_input("NRP",           h.get('nrp', ''))
        h['noDies']       = c3.text_input("No. Dies",      h.get('noDies', ''))
        h['noDoc']        = c4.text_input("No. Dokumen",   h.get('noDoc', ''))
        # Baris 3: info dokumen
        c1, _ = st.columns([2, 2])
        h['tglBerlaku']   = c1.text_input("Tgl Berlaku",   h.get('tglBerlaku', ''))
        r['header'] = h

        # ── Preview laporan dengan drag-and-drop kolom sampel ───────────────────
        st.divider()
        st.caption("↔ Geser header kolom sampel langsung di tabel untuk ubah urutan, lalu klik Simpan")
        with st.spinner("Memuat laporan..."):
            _html = self._get_html_cached(r)
        _h    = max(700, 480 + len(r.get('items', ITEMS)) * 23)
        components.html(_html, height=_h, scrolling=True)

    # ── VIEW MODE ──────────────────────────────────────────────────────────
    @st.fragment
    def _render_view(self):
        r = st.session_state.qcl_current
        if not r:
            st.session_state.qcl_view = 'list'; st.rerun()

        components.html("""<script>
setTimeout(function(){
  try{
    var d=window.parent.document;
    ['[data-testid="stMainBlockContainer"]','[data-testid="stMain"]',
     '.main','section.main','.block-container'].forEach(function(s){
      var e=d.querySelector(s); if(e) e.scrollTop=0;
    });
    window.parent.scrollTo(0,0);
    d.documentElement.scrollTop=0; d.body.scrollTop=0;
  }catch(e){}
},120);
</script>""", height=0)
        col_back, col_title, col_print = st.columns([1, 5, 1.5])
        with col_back:
            if st.button("Kembali"):
                st.session_state.qcl_view = 'list'; st.rerun()
        with col_title:
            _ng    = count_oot(r)
            _kp_ng = count_kp_ng(r)
            _parts = []
            if _ng    > 0: _parts.append(f"🔴 {_ng} NG")
            if _kp_ng > 0: _parts.append(f"⚠ {_kp_ng} KP NG")
            status = " · ".join(_parts) if _parts else "✓ Semua OK"
            st.markdown(f"**{r['header']['noPart']}** · Shift {r['header']['shift']} · {r['header']['tanggal']}  {status}")
        with col_print:
            if st.button("Print / PDF", use_container_width=True):
                with st.spinner("Menyiapkan PDF..."):
                    st.session_state['_print_rpt'] = r['id']
                st.rerun()

        n_items = len(r.get('items', ITEMS))
        html_h  = max(700, 480 + n_items * 23)
        with st.spinner("Memuat laporan..."):
            _html_v = self._get_html_cached(r)
        if st.session_state.get('_print_rpt') == r['id']:
            st.session_state.pop('_print_rpt', None)
            import time as _t
            _nonce = int(_t.time() * 1000)  # unik tiap klik → paksa re-render
            _html_v += f'<script>/* print:{_nonce} */setTimeout(function(){{window.print();}},350);</script>'
        components.html(_html_v, height=html_h, scrolling=True)

    # ── INTERNAL ───────────────────────────────────────────────────────────
    def _new_report(self):
        st.session_state.qcl_current = {
            'id': str(int(datetime.now().timestamp())),
            'createdAt': datetime.now().isoformat(),
            'submittedBy': self.current_user,
            'header': {**HDR_DEF, 'tanggal': today_str()},
            'measurements': init_meas(),
        }
        st.session_state.qcl_view = 'edit'

    def _save_current(self):
        r = st.session_state.qcl_current
        if not r:
            return
        # Persist ke local_db
        try:
            saved_id = save_report(r, self.current_user)
            if saved_id:
                r['id'] = saved_id
        except Exception:
            pass
        # Update session state
        reports = st.session_state.qcl_reports
        idx = next((i for i, x in enumerate(reports) if x['id'] == r['id']), None)
        if idx is not None:
            reports[idx] = copy.deepcopy(r)
        else:
            reports.insert(0, copy.deepcopy(r))
        st.session_state.qcl_reports = reports